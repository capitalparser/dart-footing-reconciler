"""Excel workbook reporting for validation results."""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from bs4 import BeautifulSoup
from bs4.element import Tag

MATCH_FILL = PatternFill("solid", fgColor="D9EAD3")
GAP_FILL = PatternFill("solid", fgColor="F4CCCC")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="D9E2F3"))
NUMBER_FORMAT = '#,##0;[Red](#,##0);-'


def export_validation_workbook(payload: dict[str, Any], output_path: str | Path) -> Path:
    """Write a reviewer-facing validation workbook."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    company_summary = workbook.create_sheet("Company Summary")
    note_summary = workbook.create_sheet("Note Summary")
    footing_detail = workbook.create_sheet("Footing Detail")
    gap_review = workbook.create_sheet("Gap Review")

    detail_rows = _detail_rows(payload)
    _write_dashboard(dashboard, payload)
    _write_company_summary(company_summary, payload)
    _write_note_summary(note_summary, detail_rows)
    _write_detail(footing_detail, detail_rows)
    _write_gap_review(gap_review, detail_rows)

    for sheet in workbook.worksheets:
        _freeze_and_fit(sheet)

    workbook.save(output)
    return output


def export_company_workbook(payload: dict[str, Any], output_path: str | Path) -> Path:
    """Write a single-company workbook grouped into one sheet per note number."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    note_summary = workbook.create_sheet("Note Summary")
    gap_review = workbook.create_sheet("Gap Review")

    validation_payload = _single_company_validation_payload(payload)
    detail_rows = _detail_rows(validation_payload)
    _write_company_dashboard(dashboard, payload)
    _write_note_summary(note_summary, detail_rows)
    _write_gap_review(gap_review, detail_rows)
    _write_note_sheets(workbook, detail_rows)

    for sheet in workbook.worksheets:
        _freeze_and_fit(sheet)

    workbook.save(output)
    return output


def _write_dashboard(sheet, payload: dict[str, Any]) -> None:
    summary = payload["summary"]
    rows = [
        ("Metric", "Value"),
        ("Mode", payload.get("mode") or ""),
        ("Tag", payload.get("tag") or "all"),
        ("Samples", summary.get("samples", 0)),
        ("Total tables", summary.get("total_tables", 0)),
        ("Matched", summary.get("matched", 0)),
        ("Unexplained gaps", summary.get("unexplained_gap", 0)),
        ("Match rate", _safe_rate(summary.get("matched", 0), summary.get("total_tables", 0))),
        ("Tolerance", payload.get("tolerance", "")),
        ("Manifest", payload.get("manifest", "")),
    ]
    for row in rows:
        sheet.append(row)
    sheet["B8"].number_format = "0.0%"
    _style_header(sheet, 1, 2)
    _add_status_fill(sheet, "B7")


def _write_company_dashboard(sheet, payload: dict[str, Any]) -> None:
    summary = payload["summary"]
    rows = [
        ("Metric", "Value"),
        ("Company", payload.get("company") or ""),
        ("Source", payload.get("source") or ""),
        ("Total tables", summary.get("total", 0)),
        ("Matched", summary.get("matched", 0)),
        ("Unexplained gaps", summary.get("unexplained_gap", 0)),
        ("Match rate", _safe_rate(summary.get("matched", 0), summary.get("total", 0))),
        ("Tolerance", payload.get("tolerance", "")),
    ]
    for row in rows:
        sheet.append(row)
    sheet["B7"].number_format = "0.0%"
    _style_header(sheet, 1, 2)
    _add_status_fill(sheet, "B6")


def _write_company_summary(sheet, payload: dict[str, Any]) -> None:
    headers = [
        "Sample",
        "Company",
        "Industry",
        "Status",
        "Total tables",
        "Matched",
        "Unexplained gaps",
        "Match rate",
        "Source",
    ]
    sheet.append(headers)
    for sample in payload["samples"]:
        actual = sample["actual"]
        sheet.append(
            [
                sample["name"],
                sample.get("company") or "",
                sample.get("industry") or "",
                sample["status"],
                actual.get("total", 0),
                actual.get("matched", 0),
                actual.get("unexplained_gap", 0),
                _safe_rate(actual.get("matched", 0), actual.get("total", 0)),
                sample.get("source") or "",
            ]
        )
    _style_table(sheet, "CompanySummary", len(headers))
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 8).number_format = "0.0%"
        _add_status_fill(sheet, f"G{row}")


def _write_note_summary(sheet, detail_rows: list[dict[str, Any]]) -> None:
    headers = ["Note no.", "Note title", "Tables", "Columns", "Matched", "Unexplained gaps"]
    sheet.append(headers)
    grouped: dict[tuple[str, str], dict[str, int | set[int]]] = defaultdict(
        lambda: {"tables": set(), "columns": 0, "matched": 0, "unexplained_gap": 0}
    )
    for row in detail_rows:
        key = (row["note_no"], row["note_title"])
        grouped[key]["tables"].add(row["table_index"])  # type: ignore[union-attr]
        grouped[key]["columns"] += 1  # type: ignore[operator]
        grouped[key][row["column_status"]] += 1  # type: ignore[index, operator]

    for (note_no, note_title), stats in sorted(grouped.items(), key=lambda item: item[0]):
        sheet.append(
            [
                note_no,
                note_title,
                len(stats["tables"]),
                stats["columns"],
                stats["matched"],
                stats["unexplained_gap"],
            ]
        )
    _style_table(sheet, "NoteSummary", len(headers))
    for row in range(2, sheet.max_row + 1):
        _add_status_fill(sheet, f"F{row}")


def _write_detail(
    sheet,
    detail_rows: list[dict[str, Any]],
    *,
    table_name: str = "FootingDetail",
) -> None:
    headers = [
        "Sample",
        "Company",
        "Industry",
        "Source",
        "Note no.",
        "Note title",
        "Table index",
        "Column",
        "Expected",
        "Actual",
        "Difference",
        "Column status",
        "Table status",
        "Reason",
        "Heading",
    ]
    sheet.append(headers)
    for row in detail_rows:
        sheet.append([row[key] for key in _detail_keys()])
    _style_table(sheet, table_name, len(headers))
    _format_amount_columns(sheet, (9, 10, 11))
    _apply_status_fills(sheet, 12)


def _write_note_sheets(workbook: Workbook, detail_rows: list[dict[str, Any]]) -> None:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in detail_rows:
        grouped[(row["note_no"], row["note_title"])].append(row)

    used_names = set(workbook.sheetnames)
    for (note_no, note_title), rows in sorted(grouped.items(), key=lambda item: item[0]):
        sheet = workbook.create_sheet(_note_sheet_name(note_no, note_title, used_names))
        table_name = _excel_table_name(f"NoteDetail{note_no or 'Unknown'}")
        _write_detail(sheet, rows, table_name=table_name)


def _write_gap_review(sheet, detail_rows: list[dict[str, Any]]) -> None:
    headers = [
        "Sample",
        "Company",
        "Industry",
        "Source",
        "Note no.",
        "Note title",
        "Table index",
        "Column",
        "Expected",
        "Actual",
        "Difference",
        "Column status",
        "Reviewer conclusion",
        "Reviewer memo",
        "Reason",
        "Heading",
    ]
    sheet.append(headers)
    for row in detail_rows:
        if row["column_status"] != "unexplained_gap":
            continue
        sheet.append(
            [
                row["sample"],
                row["company"],
                row["industry"],
                row["source"],
                row["note_no"],
                row["note_title"],
                row["table_index"],
                row["column"],
                row["expected"],
                row["actual"],
                row["difference"],
                row["column_status"],
                "",
                "",
                row["reason"],
                row["heading"],
            ]
        )
    _style_table(sheet, "GapReview", len(headers))
    _format_amount_columns(sheet, (9, 10, 11))
    _apply_status_fills(sheet, 12)


def _detail_rows(payload: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for sample in payload["samples"]:
        note_map = _source_note_map(sample.get("source"))
        for result in sample.get("results", []):
            note_no, note_title = _parse_note(result.get("heading", ""))
            if not note_no:
                note_no, note_title = note_map.get(result.get("table_index"), (note_no, note_title))
            for column in result.get("columns", []):
                rows.append(
                    {
                        "sample": sample["name"],
                        "company": sample.get("company") or "",
                        "industry": sample.get("industry") or "",
                        "source": sample.get("source") or "",
                        "note_no": note_no,
                        "note_title": note_title,
                        "table_index": result.get("table_index"),
                        "column": column.get("label", ""),
                        "expected": column.get("expected"),
                        "actual": column.get("actual"),
                        "difference": column.get("difference"),
                        "column_status": column.get("status", ""),
                        "table_status": result.get("status", ""),
                        "reason": result.get("reason", ""),
                        "heading": result.get("heading", ""),
                    }
                )
    return rows


def _single_company_validation_payload(payload: dict[str, Any]) -> dict[str, Any]:
    summary = payload["summary"]
    return {
        "manifest": "",
        "mode": "single-company",
        "tag": None,
        "tolerance": payload.get("tolerance", ""),
        "summary": {
            "samples": 1,
            "passed": 1,
            "failed": 0,
            "total_tables": summary.get("total", 0),
            "matched": summary.get("matched", 0),
            "unexplained_gap": summary.get("unexplained_gap", 0),
        },
        "samples": [
            {
                "name": payload.get("company") or Path(payload.get("source", "")).stem,
                "company": payload.get("company") or "",
                "industry": payload.get("industry") or "",
                "source": payload.get("source") or "",
                "status": "passed",
                "actual": summary,
                "results": payload.get("results", []),
            }
        ],
    }


def _detail_keys() -> list[str]:
    return [
        "sample",
        "company",
        "industry",
        "source",
        "note_no",
        "note_title",
        "table_index",
        "column",
        "expected",
        "actual",
        "difference",
        "column_status",
        "table_status",
        "reason",
        "heading",
    ]


def _parse_note(heading: str) -> tuple[str, str]:
    normalized = " ".join(heading.split())
    match = re.search(r"(\d+(?:-\d+)?)\.\s*([^\d]+?)(?=\s+\d+(?:-\d+)?\.|$)", normalized)
    if not match:
        return "", normalized
    return match.group(1), match.group(2).strip()


def _note_sheet_name(note_no: str, note_title: str, used_names: set[str]) -> str:
    base = f"Note {note_no}" if note_no else "Note Unknown"
    if not note_no and note_title:
        base = f"Note {note_title[:18]}"
    clean = re.sub(r"[\[\]:*?/\\]", " ", base).strip()[:31] or "Note"
    name = clean
    suffix = 2
    while name in used_names:
        suffix_text = f" {suffix}"
        name = f"{clean[: 31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used_names.add(name)
    return name


def _excel_table_name(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", value)
    if not cleaned or cleaned[0].isdigit():
        cleaned = f"Table_{cleaned}"
    return cleaned[:255]


def _source_note_map(source: str | None) -> dict[int, tuple[str, str]]:
    if not source:
        return {}
    path = Path(source)
    if not path.exists():
        return {}
    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "lxml")
    note_map: dict[int, tuple[str, str]] = {}
    for index, table in enumerate(soup.find_all("table")):
        note = _nearby_note(table)
        if note is not None:
            note_map[index] = note
    return note_map


def _nearby_note(table: Tag, max_parts: int = 12) -> tuple[str, str] | None:
    node = table.previous_sibling
    parts_seen = 0
    while node is not None and parts_seen < max_parts:
        if isinstance(node, Tag) and node.name in {"p", "div"}:
            text = " ".join(node.get_text(" ", strip=True).split())
            if text:
                note = _parse_note(text)
                if note[0]:
                    return note
                parts_seen += 1
        node = node.previous_sibling
    return None


def _safe_rate(numerator: int, denominator: int) -> float:
    if denominator == 0:
        return 0.0
    return numerator / denominator


def _style_header(sheet, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = sheet.cell(row, col)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center")


def _style_table(sheet, table_name: str, max_col: int) -> None:
    _style_header(sheet, 1, max_col)
    if sheet.max_row > 1:
        ref = f"A1:{get_column_letter(max_col)}{sheet.max_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER


def _format_amount_columns(sheet, columns: tuple[int, ...]) -> None:
    for row in range(2, sheet.max_row + 1):
        for col in columns:
            sheet.cell(row, col).number_format = NUMBER_FORMAT


def _apply_status_fills(sheet, status_col: int) -> None:
    letter = get_column_letter(status_col)
    for row in range(2, sheet.max_row + 1):
        _add_status_fill(sheet, f"{letter}{row}")


def _add_status_fill(sheet, coordinate: str) -> None:
    cell = sheet[coordinate]
    if isinstance(cell.value, int | float) and cell.value > 0:
        cell.fill = GAP_FILL
    if cell.value == "matched" or cell.value == 0:
        cell.fill = MATCH_FILL
    if cell.value == "unexplained_gap":
        cell.fill = GAP_FILL


def _freeze_and_fit(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if sheet.max_row > 1:
        sheet.auto_filter.ref = sheet.dimensions
    if sheet.title == "Gap Review":
        _add_reviewer_validation(sheet)


def _add_reviewer_validation(sheet) -> None:
    from openpyxl.worksheet.datavalidation import DataValidation

    validation = DataValidation(
        type="list",
        formula1='"OK,Explainable,Needs follow-up,Parser issue"',
        allow_blank=True,
    )
    sheet.add_data_validation(validation)
    validation.add(f"M2:M{max(sheet.max_row, 2)}")
    sheet.conditional_formatting.add(
        f"K2:K{max(sheet.max_row, 2)}",
        CellIsRule(operator="notEqual", formula=["0"], fill=GAP_FILL),
    )
