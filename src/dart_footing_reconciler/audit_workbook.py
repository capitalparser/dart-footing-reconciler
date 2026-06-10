"""Audit workpaper-style Excel workbook export."""

from __future__ import annotations

from collections import Counter
from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, quote_sheetname

from dart_footing_reconciler.checks import CheckResult
from dart_footing_reconciler.document import FullReport, ReportSection

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
CHECK_FILL = PatternFill("solid", fgColor="FFF2CC")
SOURCE_HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
MATCH_FILL = PatternFill("solid", fgColor="D9EAD3")
GAP_FILL = PatternFill("solid", fgColor="F4CCCC")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN_SIDE = Side(style="thin", color="B7B7B7")
TABLE_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
AMOUNT_FORMAT = '#,##0;[Red](#,##0);-'

VALIDATION_HEADERS = [
    "검증구분",
    "검증내용",
    "산식 / 대사흔적",
    "기준금액 / 구성항목 합계",
    "대사금액 / 표시금액",
    "차이",
    "검증결과",
    "판단근거",
    "출처",
]

SourceCellMap = dict[str, tuple[str, str]]


def export_audit_workbook(
    report: FullReport, checks: list[CheckResult], output_path: str | Path
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    fs_ws = wb.active
    fs_ws.title = "FS Summary"
    source_map: SourceCellMap = {}
    _write_fs_summary(fs_ws, report, source_map)
    summary_ws = wb.create_sheet("Validation Summary")
    _write_summary(summary_ws, report, checks)
    used_sheet_names = set(wb.sheetnames)
    for note in report.notes:
        note_ws = wb.create_sheet(_sheet_name(note, used_sheet_names))
        _write_note_sheet(
            note_ws,
            note,
            [check for check in checks if check.note_no == note.note_no],
            source_map,
        )
    for ws in wb.worksheets:
        _format_sheet(ws)
    wb.save(output)
    return output


def _write_fs_summary(ws, report: FullReport, source_map: SourceCellMap) -> None:
    ws["A1"] = "Company"
    ws["B1"] = report.company
    ws["A2"] = "Statements"
    ws["B2"] = len(report.statements)
    row = 4
    for statement in report.statements:
        ws.cell(row, 1).value = statement.title
        ws.cell(row, 1).fill = HEADER_FILL
        ws.cell(row, 1).font = Font(color="FFFFFF", bold=True)
        row += 1
        for block_index, block in enumerate(statement.blocks):
            if block.kind == "text":
                row = _write_text_block(ws, row, block.text)
                if _next_block_kind(statement.blocks, block_index) == "table":
                    row += 1
            elif block.kind == "table" and block.table is not None:
                row = _write_source_table(
                    ws,
                    row,
                    block.table.rows,
                    source_map,
                    block.location.section_id,
                    block.table.index,
                ) + 1
        row += 1


def _write_summary(ws, report: FullReport, checks: list[CheckResult]) -> None:
    ws["A1"] = "Company"
    ws["B1"] = report.company
    ws["A2"] = "Checks"
    ws["B2"] = len(checks)
    row = 4
    ws.cell(row, 1).value = "검증결과"
    ws.cell(row, 2).value = "건수"
    _style_header_row(ws, row, 1, 2)
    for status, count in Counter(check.status for check in checks).items():
        row += 1
        ws.cell(row, 1).value = _status_label(status)
        ws.cell(row, 2).value = count
        _style_body_row(ws, row, 1, 2)


def _write_note_sheet(
    ws, note: ReportSection, checks: list[CheckResult], source_map: SourceCellMap
) -> None:
    title = f"{note.note_no}. {note.title}".strip()
    ws["A1"] = title
    ws["A1"].fill = HEADER_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True)
    row = 3
    last_block_kind = ""
    for block_index, block in enumerate(note.blocks):
        if block.kind == "text":
            row = _write_text_block(ws, row, block.text)
            if _next_block_kind(note.blocks, block_index) == "table":
                row += 1
        elif block.kind == "table" and block.table is not None:
            row = _write_source_table(
                ws,
                row,
                block.table.rows,
                source_map,
                block.location.section_id,
                block.table.index,
            )
        last_block_kind = block.kind
    if last_block_kind == "text":
        row += 1
    ws.cell(row, 1).value = "검증 결과"
    ws.cell(row, 1).fill = CHECK_FILL
    ws.cell(row, 1).font = BOLD
    row += 1
    for col_idx, header in enumerate(VALIDATION_HEADERS, start=1):
        ws.cell(row, col_idx).value = header
    _style_header_row(ws, row, 1, len(VALIDATION_HEADERS))
    row += 1
    for check in checks:
        _write_check_row(ws, row, check, source_map)
        row += 1


def _sheet_name(note: ReportSection, used_sheet_names: set[str]) -> str:
    base = re.sub(r"[\[\]\:\*\?\/\\]", " ", f"Note {note.note_no}").strip() or "Note"
    base = base[:31]
    candidate = base
    suffix = 2
    while candidate in used_sheet_names:
        suffix_text = f" ({suffix})"
        candidate = f"{base[: 31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used_sheet_names.add(candidate)
    return candidate


def _write_text_block(ws, start_row: int, text: str) -> int:
    row_idx = start_row
    label = _text_label(text)
    for idx, segment in enumerate(_split_text(text)):
        label_cell = ws.cell(row_idx, 1)
        text_cell = ws.cell(row_idx, 2)
        label_cell.value = label if idx == 0 else ""
        text_cell.value = segment
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=9)
        for cell in (label_cell, text_cell):
            cell.border = TABLE_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        label_cell.fill = SOURCE_HEADER_FILL
        label_cell.font = BOLD
        row_idx += 1
    return row_idx


def _next_block_kind(blocks: list, block_index: int) -> str:
    if block_index + 1 >= len(blocks):
        return ""
    return blocks[block_index + 1].kind


def _text_label(text: str) -> str:
    stripped = text.strip()
    if stripped.startswith(("-", "ㆍ", "·", "•")):
        return "목록"
    if re.match(r"^(\d+(?:\.\d+)+|\([0-9가-힣]+\)|[①-⑳])", stripped):
        return "소제목"
    return "본문"


def _split_text(text: str, limit: int = 90) -> list[str]:
    cleaned = " ".join(text.split())
    if not cleaned:
        return []
    segments: list[str] = []
    current = ""
    for sentence in re.split(r"(?<=[.。．다])\s+", cleaned):
        if not sentence:
            continue
        if not current:
            current = sentence
        elif len(current) + 1 + len(sentence) <= limit:
            current = f"{current} {sentence}"
        else:
            segments.extend(_hard_wrap(current, limit))
            current = sentence
    if current:
        segments.extend(_hard_wrap(current, limit))
    return segments


def _hard_wrap(text: str, limit: int) -> list[str]:
    if len(text) <= limit:
        return [text]
    words = text.split()
    lines: list[str] = []
    current = ""
    for word in words:
        if not current:
            current = word
        elif len(current) + 1 + len(word) <= limit:
            current = f"{current} {word}"
        else:
            lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines


def _write_source_table(
    ws,
    start_row: int,
    rows: list[list[str]],
    source_map: SourceCellMap,
    section_id: str,
    table_index: int,
) -> int:
    row_idx = start_row
    for table_row_idx, table_row in enumerate(rows):
        for col_idx, value in enumerate(table_row, start=1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = value
            cell.border = TABLE_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if table_row_idx == 0:
                cell.fill = SOURCE_HEADER_FILL
                cell.font = BOLD
            source_map[
                f"{section_id}/table:{table_index}/row:{table_row_idx}/col:{col_idx - 1}"
            ] = (ws.title, cell.coordinate)
        row_idx += 1
    return row_idx


def _write_check_row(ws, row: int, check: CheckResult, source_map: SourceCellMap) -> None:
    formula_values = _formula_values(check, row, source_map)
    values = [
        _check_type_label(check.check_type),
        check.title,
        _trace_text(check, source_map),
        formula_values[0],
        formula_values[1],
        formula_values[2],
        _status_label(check.status),
        _reason_text(check.reason),
        _evidence_text(check, source_map),
    ]
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row, col_idx)
        cell.value = value
        cell.border = TABLE_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if col_idx in {4, 5, 6}:
            cell.number_format = AMOUNT_FORMAT
    if check.status == "matched":
        ws.cell(row, 7).fill = MATCH_FILL
    elif check.status in {"unexplained_gap", "parse_uncertain"}:
        ws.cell(row, 7).fill = GAP_FILL
    else:
        ws.cell(row, 7).fill = CHECK_FILL


def _formula_values(
    check: CheckResult, row: int, source_map: SourceCellMap
) -> tuple[int | str | None, int | str | None, str | None]:
    if check.check_type == "total_check":
        total_formula = _total_check_formula(check, source_map)
        if total_formula is not None:
            return total_formula[0], total_formula[1], f"=D{row}-E{row}"
    elif len(check.evidence) >= 2:
        left = _cell_formula(check.evidence[0].source, source_map)
        right = _cell_formula(check.evidence[1].source, source_map)
        if left and right:
            return left, right, f"=D{row}-E{row}"
    diff_formula = (
        f"=D{row}-E{row}" if check.expected is not None and check.actual is not None else None
    )
    return check.expected, check.actual, diff_formula


def _total_check_formula(
    check: CheckResult, source_map: SourceCellMap
) -> tuple[str, str] | None:
    if not check.evidence:
        return None
    source = _parse_source(check.evidence[0].source)
    if source is None:
        return None
    total_ref = _cell_formula(check.evidence[0].source, source_map)
    if total_ref is None:
        return None

    table_ref = _source_key(source["section"], source["table"], source["row"], source["col"])
    sheet_title, total_coord = source_map[table_ref]
    total_col_letter = re.match(r"([A-Z]+)", total_coord)
    total_row_number = re.search(r"(\d+)$", total_coord)
    if total_col_letter is None or total_row_number is None:
        return None

    if ":row" in check.check_id:
        first_component = _lookup_cell(
            source_map, source["section"], source["table"], source["row"], 1
        )
        last_component = _lookup_cell(
            source_map, source["section"], source["table"], source["row"], source["col"] - 1
        )
        if first_component and last_component:
            start_ref = _sheet_cell_formula(first_component[0], first_component[1])
            end_ref = _sheet_cell_formula(last_component[0], last_component[1])
            return f"=SUM({start_ref}:{end_ref.split('!', 1)[1]})", total_ref
    if ":col" in check.check_id:
        first_component = _lookup_cell(
            source_map, source["section"], source["table"], 1, source["col"]
        )
        last_component = _lookup_cell(
            source_map, source["section"], source["table"], source["row"] - 1, source["col"]
        )
        if first_component and last_component:
            start_ref = _sheet_cell_formula(first_component[0], first_component[1])
            end_ref = _sheet_cell_formula(last_component[0], last_component[1])
            return f"=SUM({start_ref}:{end_ref.split('!', 1)[1]})", total_ref
    return None


def _trace_text(check: CheckResult, source_map: SourceCellMap) -> str:
    if check.check_type == "total_check":
        source = _first_source_label(check, source_map)
        if source:
            return f"구성항목 합계 - 표시 금액 = 차이 ({source})"
        return "구성항목 합계 - 표시 금액 = 차이"
    if len(check.evidence) >= 2:
        return (
            f"{check.evidence[0].label}({_source_label(check.evidence[0].source, source_map)})"
            f" ↔ {check.evidence[1].label}({_source_label(check.evidence[1].source, source_map)})"
        )
    if check.evidence:
        return check.evidence[0].label
    if check.check_type == "prior_year_structure_change":
        return "당기 공시 구조와 전기 공시 구조 비교"
    return "검증 대상 증거 부족"


def _evidence_text(check: CheckResult, source_map: SourceCellMap) -> str:
    return " / ".join(
        f"{evidence.label}: {_source_label(evidence.source, source_map)}"
        for evidence in check.evidence
    )


def _first_source_label(check: CheckResult, source_map: SourceCellMap) -> str:
    if not check.evidence:
        return ""
    return _source_label(check.evidence[0].source, source_map)


def _source_label(source: str, source_map: SourceCellMap) -> str:
    location = source_map.get(source)
    if location is None:
        return source
    return f"{location[0]}!{location[1]}"


def _cell_formula(source: str, source_map: SourceCellMap) -> str | None:
    location = source_map.get(source)
    if location is None:
        return None
    return f"={_sheet_cell_formula(location[0], location[1])}"


def _sheet_cell_formula(sheet_title: str, coordinate: str) -> str:
    return f"{quote_sheetname(sheet_title)}!{coordinate}"


def _parse_source(source: str) -> dict[str, int | str] | None:
    match = re.match(
        r"^(?P<section>.+)/table:(?P<table>\d+)/row:(?P<row>\d+)/col:(?P<col>\d+)$",
        source,
    )
    if not match:
        return None
    return {
        "section": match.group("section"),
        "table": int(match.group("table")),
        "row": int(match.group("row")),
        "col": int(match.group("col")),
    }


def _lookup_cell(
    source_map: SourceCellMap, section: str | int, table: str | int, row: str | int, col: str | int
) -> tuple[str, str] | None:
    return source_map.get(_source_key(section, table, row, col))


def _source_key(section: str | int, table: str | int, row: str | int, col: str | int) -> str:
    return f"{section}/table:{table}/row:{row}/col:{col}"


def _reason_text(reason: str) -> str:
    labels = {
        "row total agrees": "행 구성항목 합계가 표시 금액과 일치함",
        "column total agrees": "열 구성항목 합계가 표시 금액과 일치함",
        "row total does not agree": "행 구성항목 합계와 표시 금액 간 차이가 있음",
        "column total does not agree": (
            "열 구성항목 합계와 표시 금액 간 차이가 있음"
        ),
        "no reliable total label found": (
            "합계/소계 표시를 신뢰성 있게 식별하지 못함"
        ),
        "financial statement amount agrees to note amount": (
            "재무제표 금액과 주석 금액이 일치함"
        ),
        "financial statement amount agrees within display-unit rounding": (
            "재무제표 금액과 주석 금액의 차이가 표시단위 절사 허용범위 내에 있음"
        ),
        "financial statement amount does not agree to note amount": (
            "재무제표 금액과 주석 금액 간 차이가 있음"
        ),
        "financial statement line agrees to note ending balance": (
            "재무제표 계정과 주석 기말 장부금액이 일치함"
        ),
        "financial statement line does not agree to note ending balance": (
            "재무제표 계정과 주석 기말 장부금액 간 차이가 있음"
        ),
        "cash flow statement amount agrees to note movement": (
            "현금흐름표 항목과 관련 주석 변동금액이 일치함"
        ),
        "cash flow statement amount does not agree to note movement": (
            "현금흐름표 항목과 관련 주석 변동금액 간 차이가 있음"
        ),
        "cash flow statement line agrees to note cash movement": (
            "현금흐름표 금액 크기와 주석 현금성 변동금액이 일치함"
        ),
        "cash flow statement line does not agree to note cash movement": (
            "현금흐름표 금액 크기와 주석 현금성 변동금액 간 차이가 있음"
        ),
        "current comparative amount agrees to prior current amount": (
            "당기 비교표시 전기금액과 전기 공시 당기금액이 일치함"
        ),
        "current comparative amount does not agree to prior current amount": (
            "당기 비교표시 전기금액과 전기 공시 당기금액 간 차이가 있음"
        ),
        "prior-year ending balance agrees to current-year beginning balance": (
            "전기말 주석 금액과 당기초 주석 금액이 일치함"
        ),
        "prior-year ending balance does not agree to current-year beginning balance": (
            "전기말 주석 금액과 당기초 주석 금액 간 차이가 있음"
        ),
    }
    return labels.get(reason, reason)


def _check_type_label(check_type: str) -> str:
    labels = {
        "total_check": "합계 검증 결과",
        "fs_note_match": "재무제표-주석 대사",
        "primary_balance_reconciliation": "재무제표-주석 공식 계정 대사",
        "note_rollforward_check": "주석 증감표 검산",
        "note_balance_bridge_check": "주석 잔액 연결 검증",
        "note_internal_consistency_check": "주석 내부 정합성 검증",
        "asset_note_bridge_check": "자산 주석 연결 대사",
        "note_note_match": "주석 간 대사",
        "cfs_note_match": "현금흐름표-주석 직접 대사",
        "cashflow_reconciliation": "현금흐름표-주석 현금 변동 대사",
        "prior_year_amount_match": "전기 공시 금액 대사",
        "prior_year_beginning_balance_match": "전기말-당기초 대사",
        "prior_year_structure_change": "전기 공시 구조 변경",
    }
    return labels.get(check_type, check_type)


def _status_label(status: str) -> str:
    labels = {
        "matched": "일치",
        "explainable_gap": "차이 설명 가능",
        "unexplained_gap": "미해소 차이",
        "parse_uncertain": "파싱 불확실",
        "not_tested": "검증 미수행",
    }
    return labels.get(status, status)


def _style_header_row(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.border = TABLE_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_body_row(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.border = TABLE_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _format_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    widths = {
        "A": 22,
        "B": 34,
        "C": 42,
        "D": 18,
        "E": 18,
        "F": 16,
        "G": 16,
        "H": 42,
        "I": 50,
    }
    for col_idx in range(1, max(ws.max_column, len(VALIDATION_HEADERS)) + 1):
        letter = get_column_letter(col_idx)
        if letter in widths:
            ws.column_dimensions[letter].width = widths[letter]
        else:
            ws.column_dimensions[letter].width = 16
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
