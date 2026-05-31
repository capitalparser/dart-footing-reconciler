from openpyxl import load_workbook

from dart_footing_reconciler.audit_workbook import export_audit_workbook
from dart_footing_reconciler.checks import CheckEvidence, CheckResult, MATCHED
from dart_footing_reconciler.document import (
    FullReport,
    ReportBlock,
    ReportSection,
    ReportTable,
    SourceLocation,
)


def test_export_audit_workbook_renders_note_then_validation_block(tmp_path):
    table = ReportTable(
        0, [["구분", "합계"], ["기초", "100"]], "11. 유형자산", SourceLocation("note:11", 1, 0)
    )
    note = ReportSection(
        section_id="note:11",
        title="유형자산",
        kind="note",
        note_no="11",
        blocks=[
            ReportBlock("text", "유형자산 변동내역입니다.", None, SourceLocation("note:11", 0)),
            ReportBlock("table", "", table, SourceLocation("note:11", 1, 0)),
        ],
    )
    report = FullReport(str(tmp_path / "report.html"), "Sample Co", [], [note])
    checks = [
        CheckResult(
            "total:11:0",
            "total_check",
            MATCHED,
            "note",
            "11",
            "row total",
            100,
            100,
            0,
            1,
            "row total agrees",
            [CheckEvidence("기초", 100, "note:11/table:0/row:1/col:1")],
        )
    ]
    output = tmp_path / "workpaper.xlsx"

    export_audit_workbook(report, checks, output)

    wb = load_workbook(output)
    assert "FS Summary" in wb.sheetnames
    assert "Validation Summary" in wb.sheetnames
    ws = wb["Note 11"]
    assert ws["A1"].value == "11. 유형자산"
    assert ws["A3"].value == "본문"
    assert ws["B3"].value == "유형자산 변동내역입니다."
    assert ws["A7"].value == "검증 결과"
    assert ws["A8"].value == "검증구분"
    assert ws["B8"].value == "검증내용"
    assert ws["C8"].value == "산식 / 대사흔적"
    assert ws["D8"].value == "기준금액 / 구성항목 합계"
    assert ws["E8"].value == "대사금액 / 표시금액"
    assert ws["F8"].value == "차이"
    assert ws["G8"].value == "검증결과"
    assert ws["A9"].value == "합계 검증 결과"
    assert ws["C9"].value == "구성항목 합계 - 표시 금액 = 차이 (Note 11!B6)"
    assert ws["F9"].value == "=D9-E9"
    assert ws["G9"].value == "일치"
    assert ws["A8"].fill.fgColor.rgb == "001F4E78"
    assert ws["A9"].border.left.style == "thin"


def test_export_audit_workbook_renders_statement_source_tables(tmp_path):
    statement_table = ReportTable(
        0,
        [["구분", "당기"], ["자산총계", "1,000"]],
        "재무상태표",
        SourceLocation("statement:재무상태표", 0, 0),
    )
    statement = ReportSection(
        section_id="statement:재무상태표",
        title="재무상태표",
        kind="statement",
        note_no="",
        blocks=[ReportBlock("table", "", statement_table, SourceLocation("statement:재무상태표", 0, 0))],
    )
    report = FullReport(str(tmp_path / "report.html"), "Sample Co", [statement], [])
    output = tmp_path / "workpaper.xlsx"

    export_audit_workbook(report, [], output)

    ws = load_workbook(output)["FS Summary"]
    values = [cell.value for row in ws.iter_rows() for cell in row]
    assert "재무상태표" in values
    assert "자산총계" in values
    assert "1,000" in values


def test_export_audit_workbook_uses_business_labels_for_matching_checks(tmp_path):
    note = ReportSection(
        section_id="note:11",
        title="유형자산",
        kind="note",
        note_no="11",
        blocks=[ReportBlock("text", "유형자산 내용입니다.", None, SourceLocation("note:11", 0))],
    )
    report = FullReport(str(tmp_path / "report.html"), "Sample Co", [], [note])
    checks = [
        CheckResult(
            "fs_note:ppe:11",
            "fs_note_match",
            MATCHED,
            "report",
            "11",
            "유형자산 FS to note match",
            1000,
            1000,
            0,
            1,
            "financial statement amount agrees to note amount",
            [
                CheckEvidence("재무상태표 유형자산", 1000, "statement/table:0/row:1/col:1"),
                CheckEvidence("주석 유형자산 장부금액", 1000, "note:11/table:1/row:2/col:3"),
            ],
        )
    ]
    output = tmp_path / "workpaper.xlsx"

    export_audit_workbook(report, checks, output)

    ws = load_workbook(output, data_only=False)["Note 11"]
    assert ws["A5"].value == "검증 결과"
    assert ws["A6"].value == "검증구분"
    assert ws["A7"].value == "재무제표-주석 대사"
    assert ws["C7"].value == (
        "재무상태표 유형자산(statement/table:0/row:1/col:1)"
        " ↔ 주석 유형자산 장부금액(note:11/table:1/row:2/col:3)"
    )
    assert ws["D7"].value == 1000
    assert ws["E7"].value == 1000
    assert ws["F7"].value == "=D7-E7"
    assert ws["G7"].value == "일치"


def test_export_audit_workbook_splits_long_text_across_rows(tmp_path):
    long_text = "문장 하나입니다. " * 20
    note = ReportSection(
        section_id="note:2",
        title="중요한 회계정책",
        kind="note",
        note_no="2",
        blocks=[ReportBlock("text", long_text, None, SourceLocation("note:2", 0))],
    )
    report = FullReport(str(tmp_path / "report.html"), "Sample Co", [], [note])
    output = tmp_path / "workpaper.xlsx"

    export_audit_workbook(report, [], output)

    ws = load_workbook(output)["Note 2"]
    assert ws["A3"].value == "본문"
    assert ws["B3"].value != long_text
    assert ws["B4"].value
    assert len(ws["B3"].value) <= 100
    assert "B3:I3" in ws.merged_cells
    assert "B4:I4" in ws.merged_cells


def test_export_audit_workbook_uses_unique_duplicate_note_sheet_names(tmp_path):
    notes = [
        ReportSection("note:1:a", "일반사항 (연결)", "note", "1", []),
        ReportSection("note:1:b", "일반사항", "note", "1", []),
    ]
    report = FullReport(str(tmp_path / "report.html"), "Sample Co", [], notes)
    output = tmp_path / "workpaper.xlsx"

    export_audit_workbook(report, [], output)

    workbook = load_workbook(output)
    assert "Note 1" in workbook.sheetnames
    assert "Note 1 (2)" in workbook.sheetnames
    assert "Note 11" not in workbook.sheetnames


def test_export_audit_workbook_references_source_cells_in_total_check(tmp_path):
    table = ReportTable(
        0,
        [["구분", "토지", "건물", "합계"], ["기초", "100", "200", "300"]],
        "11. 유형자산",
        SourceLocation("note:11", 0, 0),
    )
    note = ReportSection(
        section_id="note:11",
        title="유형자산",
        kind="note",
        note_no="11",
        blocks=[ReportBlock("table", "", table, SourceLocation("note:11", 0, 0))],
    )
    report = FullReport(str(tmp_path / "report.html"), "Sample Co", [], [note])
    checks = [
        CheckResult(
            "total:11:table0:row1",
            "total_check",
            MATCHED,
            "note",
            "11",
            "기초 row total",
            300,
            300,
            0,
            0,
            "row total agrees",
            [CheckEvidence("기초", 300, "note:11/table:0/row:1/col:3")],
        )
    ]
    output = tmp_path / "workpaper.xlsx"

    export_audit_workbook(report, checks, output)

    ws = load_workbook(output, data_only=False)["Note 11"]
    assert ws["D7"].value == "=SUM('Note 11'!B4:C4)"
    assert ws["E7"].value == "='Note 11'!D4"
    assert ws["F7"].value == "=D7-E7"
    assert "Note 11!D4" in ws["I7"].value
