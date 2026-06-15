import json

from openpyxl import load_workbook
from typer.testing import CliRunner

from dart_footing_reconciler.cli import app


def test_cli_foot_outputs_json(tmp_path) -> None:
    source = tmp_path / "report.html"
    source.write_text(
        """
        <p>15. 무형자산</p>
        <p>(2) 당기와 전기 중 무형자산의 변동내용은 다음과 같습니다.</p>
        <table>
          <tr><th>구분</th><th>합계</th></tr>
          <tr><td>기초</td><td>1,000</td></tr>
          <tr><td>취득</td><td>250</td></tr>
          <tr><td>상각비</td><td>100</td></tr>
          <tr><td>기말</td><td>1,150</td></tr>
        </table>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["foot", str(source), "--format", "json"])

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert payload["summary"]["total"] == 1
    assert payload["summary"]["matched"] == 1
    assert payload["results"][0]["columns"][0]["difference"] == 0


def test_cli_foot_accepts_local_dsd_with_korean_encoding(tmp_path) -> None:
    source = tmp_path / "report.dsd"
    source.write_bytes(
        """
        <DOCUMENT>
        <p>14. 유형자산</p>
        <p>당기 중 유형자산의 변동내용은 다음과 같습니다.</p>
        <table>
          <tr><th>구분</th><th>합계</th></tr>
          <tr><td>기초</td><td>1,000</td></tr>
          <tr><td>취득</td><td>250</td></tr>
          <tr><td>감가상각비</td><td>100</td></tr>
          <tr><td>기말</td><td>1,150</td></tr>
        </table>
        </DOCUMENT>
        """.encode("cp949")
    )

    result = CliRunner().invoke(app, ["foot", str(source), "--format", "json"])

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert payload["input_format"] == "dsd"
    assert payload["summary"]["matched"] == 1
    assert payload["results"][0]["status"] == "matched"


def test_cli_foot_rejects_pdf_until_pdf_table_extraction_is_supported(tmp_path) -> None:
    source = tmp_path / "report.pdf"
    source.write_bytes(b"%PDF-1.7\n")

    result = CliRunner().invoke(app, ["foot", str(source), "--format", "json"])

    assert result.exit_code != 0
    assert "PDF footing is not supported" in result.output
    assert "DSD or HTML" in result.output


def test_cli_foot_rejects_network_sources() -> None:
    result = CliRunner().invoke(app, ["foot", "https://dart.fss.or.kr/report.dsd"])

    assert result.exit_code != 0
    assert "local file path" in result.output


def test_cli_foot_excel_outputs_company_note_workbook(tmp_path) -> None:
    source = tmp_path / "report.html"
    source.write_text(
        """
        <p>11. 유형자산</p>
        <p>유형자산의 변동내용은 다음과 같습니다.</p>
        <table>
          <tr><th>구분</th><th>합계</th></tr>
          <tr><td>기초</td><td>1,000</td></tr>
          <tr><td>취득</td><td>250</td></tr>
          <tr><td>감가상각비</td><td>100</td></tr>
          <tr><td>기말</td><td>1,150</td></tr>
        </table>
        """,
        encoding="utf-8",
    )
    output = tmp_path / "company_review.xlsx"

    result = CliRunner().invoke(
        app,
        ["foot-excel", str(source), str(output), "--company", "Sample Co"],
    )

    assert result.exit_code == 0
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["Dashboard", "Note Summary", "Gap Review", "Note 11"]
    assert workbook["Dashboard"]["B2"].value == "Sample Co"
    assert workbook["Note 11"]["E2"].value == "11"


def test_cli_coverage_report_outputs_full_note_counts(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <html><body>
          <p>1. 일반사항</p>
          <table><tr><th>구분</th><th>내용</th></tr><tr><td>회사</td><td>샘플</td></tr></table>
          <p>11. 유형자산</p>
          <table><tr><th>구분</th><th>합계</th></tr><tr><td>기말</td><td>1,000</td></tr></table>
        </body></html>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["coverage-report", str(source), "--company", "Sample Co"])

    assert result.exit_code == 0
    assert "company: Sample Co" in result.output
    assert "total_notes: 2" in result.output
    assert "total_tables: 2" in result.output
    assert "unknown_layout_tables:" in result.output


def test_cli_candidate_report_outputs_extraction_summary(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <html><body>
          <p>11. 유형자산</p>
          <table>
            <tr><th>구분</th><th>합계</th></tr>
            <tr><td>기초</td><td>100</td></tr>
            <tr><td>취득</td><td>50</td></tr>
            <tr><td>기말</td><td>150</td></tr>
          </table>
          <p>12. 유형자산</p>
          <table>
            <tr><th>구분</th><th>내용</th></tr>
            <tr><td>취득</td><td>50</td></tr>
          </table>
        </body></html>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["candidate-report", str(source), "--company", "Sample Co"])

    assert result.exit_code == 0
    assert "company: Sample Co" in result.output
    assert "total_note_tables: 2" in result.output
    assert "orientation row_oriented: 1" in result.output
    assert "validation_relevant_unknown_layout_items: 1" in result.output
    assert "validation_relevance asset_rollforward_candidate: 1" in result.output
    assert "verification_candidates: 3" in result.output
    assert "verification_formulas: 1" in result.output
    assert "matched_formulas: 1" in result.output


def test_cli_candidate_report_counts_component_net_formula(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <html><body>
          <p>7. 무형자산</p>
          <table>
            <tr><th>계 정 과 목</th><th>취득원가</th><th>상각누계액</th><th>손상차손누계액</th><th>기 말</th></tr>
            <tr><td>회원권</td><td>1,405,264</td><td>-</td><td>(435,261)</td><td>970,003</td></tr>
            <tr><td>소프트웨어</td><td>1,238,823</td><td>(1,178,732)</td><td>-</td><td>60,091</td></tr>
            <tr><td>합 계</td><td>2,644,087</td><td>(1,178,732)</td><td>(435,261)</td><td>1,030,094</td></tr>
          </table>
        </body></html>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["candidate-report", str(source), "--company", "Sample Co"])

    assert result.exit_code == 0
    assert "layout asset_cost_accumulated_summary: 1" in result.output
    assert "verification_candidates: 4" in result.output
    assert "verification_formulas: 1" in result.output
    assert "matched_formulas: 1" in result.output


def test_cli_candidate_report_counts_debt_split_formula(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <html><body>
          <p>15. 차입금 및 사채</p>
          <table>
            <tr><th></th><th>차입금명칭</th><th>차입금명칭</th><th>차입금명칭 합계</th></tr>
            <tr><td>만기일</td><td></td><td>2029-06-15</td><td></td></tr>
            <tr><td>연이자율</td><td>0.0175</td><td></td><td></td></tr>
            <tr><td>차입금</td><td></td><td>6,294,460</td><td>6,294,460</td></tr>
            <tr><td>1년이내 만기도래분</td><td></td><td></td><td>(1,223,560)</td></tr>
            <tr><td>비유동성 차입금</td><td></td><td></td><td>5,070,900</td></tr>
          </table>
        </body></html>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["candidate-report", str(source), "--company", "Sample Co"])

    assert result.exit_code == 0
    assert "layout debt_instrument_detail_summary: 1" in result.output
    assert "verification_candidates: 3" in result.output
    assert "verification_formulas: 1" in result.output
    assert "matched_formulas: 1" in result.output


def test_cli_candidate_report_counts_expense_summary_formula(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <html><body>
          <p>26. 판매비와 관리비</p>
          <table>
            <tr><th></th><th>금액</th></tr>
            <tr><td>급여, 판관비</td><td>100</td></tr>
            <tr><td>감가상각비, 판관비</td><td>30</td></tr>
            <tr><td>기타판매비와관리비</td><td>20</td></tr>
            <tr><td>합계</td><td>150</td></tr>
          </table>
        </body></html>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["candidate-report", str(source), "--company", "Sample Co"])

    assert result.exit_code == 0
    assert "layout selling_admin_expense_summary: 1" in result.output
    assert "verification_candidates: 4" in result.output
    assert "verification_formulas: 1" in result.output
    assert "matched_formulas: 1" in result.output


def test_cli_candidate_report_counts_operating_expense_summary_formula(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <html><body>
          <p>22. 영업비용(별도)</p>
          <table>
            <tr><th></th><th>금액</th></tr>
            <tr><td>가스매출원가</td><td>18,228,713</td></tr>
            <tr><td>금융매출원가</td><td>27,734,813</td></tr>
            <tr><td>급여, 판관비</td><td>3,409,843</td></tr>
            <tr><td>감가상각비, 판관비</td><td>1,290,295</td></tr>
            <tr><td>기타판매비와관리비</td><td>437,513</td></tr>
            <tr><td>합계</td><td>51,101,178</td></tr>
          </table>
        </body></html>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["candidate-report", str(source), "--company", "Sample Co"])

    assert result.exit_code == 0
    assert "layout operating_expense_summary: 1" in result.output
    assert "verification_candidates: 6" in result.output
    assert "verification_formulas: 1" in result.output
    assert "matched_formulas: 1" in result.output


def test_cli_candidate_report_counts_single_row_functional_allocation_candidate(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <html><body>
          <p>10. 유형자산(별도)</p>
          <p>감가상각비의 기능별 배분 당기</p>
          <table>
            <tr><th></th><th></th><th>감가상각비, 유형자산</th></tr>
            <tr><td>기능별 항목</td><td>영업비용</td><td>580,870</td></tr>
          </table>
        </body></html>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["candidate-report", str(source), "--company", "Sample Co"])

    assert result.exit_code == 0
    assert "layout functional_expense_single_row_allocation: 1" in result.output
    assert "verification_candidates: 1" in result.output
    assert "verification_formulas: 0" in result.output
    assert "matched_formulas: 0" in result.output


def test_cli_candidate_report_counts_net_debt_bridge_formulas(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <html><body>
          <p>29. 영업으로부터 창출된 현금</p>
          <p>재무활동에서 생기는 부채의 조정</p>
          <table>
            <tr><th></th><th>유동성사채</th><th>리스 부채</th><th>장기 차입금</th></tr>
            <tr><td>기초 순부채</td><td>100</td><td>50</td><td>70</td></tr>
            <tr><td>현금흐름</td><td>(20)</td><td>(10)</td><td>30</td></tr>
            <tr><td>이자비용</td><td>5</td><td>3</td><td>0</td></tr>
            <tr><td>기말 순부채</td><td>85</td><td>43</td><td>100</td></tr>
          </table>
        </body></html>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["candidate-report", str(source), "--company", "Sample Co"])

    assert result.exit_code == 0
    assert "layout net_debt_bridge: 1" in result.output
    assert "verification_candidates: 12" in result.output
    assert "verification_formulas: 3" in result.output
    assert "matched_formulas: 3" in result.output


def test_cli_candidate_report_counts_financing_debt_bridge_formulas(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <html><body>
          <p>연결재무제표 주석</p>
          <p>37. 현금흐름표</p>
          <p>재무활동에서 생기는 부채의 조정에 관한 공시 (단위 : 천원)</p>
          <table>
            <tr><th></th><th>단기차입금</th><th>장기차입금</th><th>사채</th><th>리스부채</th><th>미지급배당금</th></tr>
            <tr><td>재무활동에서 생기는 기초 부채</td><td>19,135</td><td>117,608</td><td>146,640</td><td>6,517</td><td>3</td></tr>
            <tr><td>차입금의 증가, 재무활동에서 생기는 부채</td><td>104,568</td><td>41,236</td><td>0</td><td>0</td><td>0</td></tr>
            <tr><td>차입금의 감소, 재무활동에서 생기는 부채</td><td>(54,502)</td><td>(37,159)</td><td>(78,010)</td><td>(2,308)</td><td>0</td></tr>
            <tr><td>그 밖의 변동, 재무활동에서 생기는 부채의 증가(감소)</td><td>0</td><td>283</td><td>1,321</td><td>2,104</td><td>23,631</td></tr>
            <tr><td>재무활동에서 생기는 기말 부채</td><td>69,201</td><td>121,968</td><td>69,951</td><td>6,313</td><td>23,634</td></tr>
          </table>
        </body></html>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["candidate-report", str(source), "--company", "Sample Co"])

    assert result.exit_code == 0
    assert "layout net_debt_bridge: 1" in result.output
    assert "verification_candidates: 25" in result.output
    assert "verification_formulas: 5" in result.output
    assert "matched_formulas: 5" in result.output


def test_cli_candidate_report_counts_defined_benefit_rollforward_formulas(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <html><body>
          <p>13. 순확정급여부채(자산)</p>
          <table>
            <tr><th></th><th></th><th>확정급여채무의 현재가치</th><th>사외적립자산</th></tr>
            <tr><td>기초금액</td><td>기초금액</td><td>100</td><td>70</td></tr>
            <tr><td>당기근무원가</td><td>당기근무원가</td><td>30</td><td>0</td></tr>
            <tr><td>이자비용(수익)</td><td>이자비용(수익)</td><td>5</td><td>(3)</td></tr>
            <tr><td>재측정요소:</td><td>재측정요소:</td><td></td><td></td></tr>
            <tr><td>재측정요소:</td><td>총 재측정손익</td><td>(10)</td><td>(5)</td></tr>
            <tr><td>기말금액</td><td>기말금액</td><td>125</td><td>62</td></tr>
          </table>
        </body></html>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["candidate-report", str(source), "--company", "Sample Co"])

    assert result.exit_code == 0
    assert "layout defined_benefit_rollforward: 1" in result.output
    assert "verification_candidates: 10" in result.output
    assert "verification_formulas: 2" in result.output
    assert "matched_formulas: 2" in result.output


def test_cli_candidate_report_counts_defined_benefit_formulas_from_employee_benefit_title(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <html><body>
          <p>18. 퇴직급여제도</p>
          <table>
            <tr><th></th><th></th><th>확정급여채무의 현재가치</th><th>사외적립자산</th></tr>
            <tr><td>기초</td><td>기초</td><td>100</td><td>70</td></tr>
            <tr><td>당기근무원가</td><td>당기근무원가</td><td>30</td><td>0</td></tr>
            <tr><td>이자비용(이자수익)</td><td>이자비용(이자수익)</td><td>5</td><td>(3)</td></tr>
            <tr><td>기말</td><td>기말</td><td>135</td><td>67</td></tr>
          </table>
        </body></html>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["candidate-report", str(source), "--company", "Sample Co"])

    assert result.exit_code == 0
    assert "layout defined_benefit_rollforward: 1" in result.output
    assert "verification_candidates: 8" in result.output
    assert "verification_formulas: 2" in result.output
    assert "matched_formulas: 2" in result.output


def test_cli_candidate_report_counts_inventory_allowance_rollforward_formula(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <html><body>
          <p>6. 재고자산</p>
          <p>재고자산 평가충당금의 변동내역 당기 (단위 : 천원)</p>
          <table>
            <tr><th></th><th>재고자산 평가충당금</th></tr>
            <tr><td>기초재고자산</td><td>(3,969,701)</td></tr>
            <tr><td>재고자산 평가손실환입</td><td>427,066</td></tr>
            <tr><td>재고자산 평가손실</td><td>(961,511)</td></tr>
            <tr><td>재고자산 폐기</td><td>0</td></tr>
            <tr><td>기타 (주1)</td><td>(1,553)</td></tr>
            <tr><td>기말재고자산</td><td>(4,505,699)</td></tr>
          </table>
        </body></html>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["candidate-report", str(source), "--company", "Sample Co"])

    assert result.exit_code == 0
    assert "layout inventory_allowance_rollforward: 1" in result.output
    assert "verification_candidates: 6" in result.output
    assert "verification_formulas: 1" in result.output
    assert "matched_formulas: 1" in result.output


def test_cli_candidate_report_counts_row_oriented_provision_rollforward_formulas(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <html><body>
          <p>18. 충당부채</p>
          <p>기타충당부채의 변동 당기 (단위 : 천원)</p>
          <table>
            <tr><th></th><th></th><th>사후처리, 복구, 정화 비용을 위한 충당부채</th><th>기타장기종업원급여부채</th><th>기타충당부채 합계</th></tr>
            <tr><td>기초 기타충당부채</td><td>기초 기타충당부채</td><td>778,101</td><td>1,697,795</td><td>2,475,896</td></tr>
            <tr><td>기타충당부채의 변동에 대한 조정</td><td>당기에 추가된 충당부채 합계, 기타충당부채</td><td>279,165</td><td>312,625</td><td>591,790</td></tr>
            <tr><td>기타충당부채의 변동에 대한 조정</td><td>사용된 충당부채, 기타충당부채</td><td>0</td><td>(150,000)</td><td>(150,000)</td></tr>
            <tr><td>기말 기타충당부채</td><td>기말 기타충당부채</td><td>1,057,266</td><td>1,860,420</td><td>2,917,686</td></tr>
          </table>
        </body></html>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["candidate-report", str(source), "--company", "Sample Co"])

    assert result.exit_code == 0
    assert "layout provision_rollforward: 1" in result.output
    assert "verification_candidates: 12" in result.output
    assert "verification_formulas: 3" in result.output
    assert "matched_formulas: 3" in result.output


def test_cli_candidate_report_counts_credit_risk_exposure_formula(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <html><body>
          <p>31. 재무위험관리</p>
          <p>신용위험 익스포저에 대한 공시</p>
          <table>
            <tr><th></th><th>신용위험</th></tr>
            <tr><td>현금성자산</td><td>100</td></tr>
            <tr><td>단기당기손익-공정가치측정금융자산</td><td>200</td></tr>
            <tr><td>매출채권</td><td>300</td></tr>
            <tr><td>기타비유동금융자산</td><td>50</td></tr>
            <tr><td>합계</td><td>650</td></tr>
          </table>
        </body></html>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["candidate-report", str(source), "--company", "Sample Co"])

    assert result.exit_code == 0
    assert "layout credit_risk_exposure_summary: 1" in result.output
    assert "verification_candidates: 5" in result.output
    assert "verification_formulas: 1" in result.output
    assert "matched_formulas: 1" in result.output


def test_cli_candidate_report_counts_liquidity_maturity_formulas(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <html><body>
          <p>31. 재무위험관리</p>
          <p>유동성위험 관리 목적으로 보유한 금융자산에 대한 만기분석 공시</p>
          <table>
            <tr><th></th><th>3개월 이내</th><th>3개월 초과 1년 이내</th><th>1년 초과 2년 이내</th><th>2년 초과</th><th>합계 구간 합계</th></tr>
            <tr><td>차입금 및 사채</td><td>10</td><td>20</td><td>30</td><td>40</td><td>100</td></tr>
            <tr><td>리스부채</td><td>1</td><td>2</td><td>3</td><td>4</td><td>10</td></tr>
            <tr><td>합계</td><td>11</td><td>22</td><td>33</td><td>44</td><td>110</td></tr>
          </table>
        </body></html>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["candidate-report", str(source), "--company", "Sample Co"])

    assert result.exit_code == 0
    assert "layout liquidity_maturity_analysis: 1" in result.output
    assert "verification_candidates: 15" in result.output
    assert "verification_formulas: 3" in result.output
    assert "matched_formulas: 3" in result.output


def test_cli_candidate_report_counts_lease_expense_formula(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <html><body>
          <p>34. 리스</p>
          <p>리스 관련 비용</p>
          <table>
            <tr><th></th><th>자산</th><th>자산</th><th>자산 합계</th></tr>
            <tr><td></td><td>사용권자산</td><td>사용권자산</td><td>자산 합계</td></tr>
            <tr><td></td><td>부동산</td><td>차량운반구</td><td>자산 합계</td></tr>
            <tr><td>감가상각비, 사용권자산</td><td>100</td><td>20</td><td>120</td></tr>
            <tr><td>리스부채에 대한 이자비용(금융비용에 포함)</td><td></td><td></td><td>30</td></tr>
            <tr><td>단기리스료</td><td></td><td></td><td>4</td></tr>
          </table>
        </body></html>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["candidate-report", str(source), "--company", "Sample Co"])

    assert result.exit_code == 0
    assert "layout lease_expense_summary: 1" in result.output
    assert "verification_candidates: 5" in result.output
    assert "verification_formulas: 1" in result.output
    assert "matched_formulas: 1" in result.output


def test_cli_candidate_report_counts_discontinued_operation_income_formulas(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <html><body>
          <p>35. 매각예정처분자산(부채)집단과 중단영업</p>
          <table>
            <tr><th></th><th></th><th>중단영업</th></tr>
            <tr><td>매출액</td><td>매출액</td><td>86,773,808</td></tr>
            <tr><td>매출원가</td><td>매출원가</td><td>71,401,947</td></tr>
            <tr><td>매출총이익</td><td>매출총이익</td><td>15,371,861</td></tr>
            <tr><td>판매비와관리비</td><td>판매비와관리비</td><td>16,214,619</td></tr>
            <tr><td>영업이익(손실)</td><td>영업이익(손실)</td><td>(842,758)</td></tr>
            <tr><td>기타이익</td><td>기타이익</td><td>738,185</td></tr>
            <tr><td>기타손실</td><td>기타손실</td><td>129,922</td></tr>
            <tr><td>금융수익</td><td>금융수익</td><td>555,985</td></tr>
            <tr><td>금융비용</td><td>금융비용</td><td>2,158,514</td></tr>
            <tr><td>법인세비용차감전순이익(손실)</td><td>법인세비용차감전순이익(손실)</td><td>(1,837,024)</td></tr>
            <tr><td>중단영업 법인세비용(수익)</td><td>중단영업 법인세비용(수익)</td><td>137,049</td></tr>
            <tr><td>중단영업이익(손실)</td><td>중단영업이익(손실)</td><td>(1,974,073)</td></tr>
            <tr><td>중단영업처분이익</td><td>중단영업처분이익</td><td>21,874,959</td></tr>
            <tr><td>중단영업순이익</td><td>중단영업순이익</td><td>19,900,886</td></tr>
            <tr><td>중단영업순이익</td><td>지배기업의 소유주에게 귀속될 중단영업손익</td><td>20,368,366</td></tr>
            <tr><td>중단영업순이익</td><td>비지배지분에 귀속될 중단영업이익(손실)</td><td>(467,480)</td></tr>
          </table>
        </body></html>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["candidate-report", str(source), "--company", "Sample Co"])

    assert result.exit_code == 0
    assert "layout discontinued_operation_income_statement: 1" in result.output
    assert "verification_candidates: 16" in result.output
    assert "verification_formulas: 6" in result.output
    assert "matched_formulas: 6" in result.output


def test_cli_candidate_report_counts_discontinued_operation_cashflow_formula(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <html><body>
          <p>35. 매각예정처분자산(부채)집단과 중단영업</p>
          <table>
            <tr><th></th><th>중단영업</th></tr>
            <tr><td>중단영업영업활동현금흐름</td><td>10</td></tr>
            <tr><td>중단영업투자활동현금흐름</td><td>(3)</td></tr>
            <tr><td>중단영업재무활동현금흐름</td><td>2</td></tr>
            <tr><td>합계</td><td>9</td></tr>
          </table>
        </body></html>
        """,
        encoding="utf-8",
    )

    result = CliRunner().invoke(app, ["candidate-report", str(source), "--company", "Sample Co"])

    assert result.exit_code == 0
    assert "layout discontinued_operation_cashflow_summary: 1" in result.output
    assert "verification_candidates: 4" in result.output
    assert "verification_formulas: 1" in result.output
    assert "matched_formulas: 1" in result.output
