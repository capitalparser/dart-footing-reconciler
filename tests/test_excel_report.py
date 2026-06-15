from openpyxl import load_workbook

from dart_footing_reconciler.excel import export_company_workbook, export_validation_workbook


def test_export_validation_workbook_writes_dashboard_and_note_detail(tmp_path) -> None:
    payload = {
        "manifest": "manifest.json",
        "mode": "conservative",
        "tag": "manufacturing",
        "tolerance": 1,
        "summary": {
            "samples": 1,
            "passed": 1,
            "failed": 0,
            "total_tables": 1,
            "matched": 0,
            "unexplained_gap": 1,
        },
        "samples": [
            {
                "name": "sample-manufacturer",
                "company": "Sample Manufacturing",
                "industry": "manufacturing",
                "source": "sample.html",
                "status": "passed",
                "actual": {"total": 1, "matched": 0, "unexplained_gap": 1},
                "results": [
                    {
                        "table_index": 12,
                        "heading": "11. 유형자산 (연결)",
                        "status": "unexplained_gap",
                        "reason": "one or more columns do not foot",
                        "columns": [
                            {
                                "label": "기계장치",
                                "expected": 1000,
                                "actual": 980,
                                "difference": -20,
                                "status": "unexplained_gap",
                            }
                        ],
                    }
                ],
            }
        ],
    }
    output = tmp_path / "footing_review.xlsx"

    export_validation_workbook(payload, output)

    workbook = load_workbook(output)
    assert workbook.sheetnames == [
        "Dashboard",
        "Company Summary",
        "Note Summary",
        "Footing Detail",
        "Gap Review",
    ]
    assert workbook["Dashboard"]["A1"].value == "항목"
    assert workbook["Dashboard"]["A5"].value == "전체 대사 항목"
    assert workbook["Dashboard"]["A6"].value == "일치"
    assert workbook["Dashboard"]["A7"].value == "미해소 차이"
    assert workbook["Dashboard"]["B4"].value == 1
    assert workbook["Company Summary"]["A2"].value == "sample-manufacturer"
    assert workbook["Note Summary"]["A2"].value == "11"
    assert workbook["Note Summary"]["B2"].value == "유형자산 (연결)"
    assert workbook["Footing Detail"]["E2"].value == "11"
    assert workbook["Gap Review"]["K2"].value == -20


def test_export_validation_workbook_backfills_note_from_source_html(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <p>11. 유형자산</p>
        <p>(1) 당기 및 전기 중 유형자산 장부금액의 변동내역은 다음과 같습니다.</p>
        <p>(당기)</p>
        <table>
          <tr><th>구분</th><th>합계</th></tr>
          <tr><td>기초</td><td>1,000</td></tr>
          <tr><td>기말</td><td>1,000</td></tr>
        </table>
        """,
        encoding="utf-8",
    )
    payload = {
        "manifest": "manifest.json",
        "mode": "conservative",
        "tag": None,
        "tolerance": 1,
        "summary": {
            "samples": 1,
            "passed": 1,
            "failed": 0,
            "total_tables": 1,
            "matched": 1,
            "unexplained_gap": 0,
        },
        "samples": [
            {
                "name": "sample",
                "source": str(source),
                "status": "passed",
                "actual": {"total": 1, "matched": 1, "unexplained_gap": 0},
                "results": [
                    {
                        "table_index": 0,
                        "heading": "(1) 당기 및 전기 중 유형자산 장부금액의 변동내역은 다음과 같습니다. (당기)",
                        "status": "matched",
                        "reason": "all columns foot within tolerance",
                        "columns": [
                            {
                                "label": "합계",
                                "expected": 1000,
                                "actual": 1000,
                                "difference": 0,
                                "status": "matched",
                            }
                        ],
                    }
                ],
            }
        ],
    }
    output = tmp_path / "footing_review.xlsx"

    export_validation_workbook(payload, output)

    workbook = load_workbook(output)
    assert workbook["Footing Detail"]["E2"].value == "11"
    assert workbook["Footing Detail"]["F2"].value == "유형자산"


def test_export_company_workbook_creates_one_sheet_per_note(tmp_path) -> None:
    source = tmp_path / "sample.html"
    source.write_text(
        """
        <p>11. 유형자산</p>
        <p>(1) 유형자산 변동내역입니다.</p>
        <table>
          <tr><th>구분</th><th>합계</th></tr>
          <tr><td>기초</td><td>1,000</td></tr>
          <tr><td>기말</td><td>1,000</td></tr>
        </table>
        <p>12. 무형자산</p>
        <p>(1) 무형자산 변동내역입니다.</p>
        <table>
          <tr><th>구분</th><th>합계</th></tr>
          <tr><td>기초</td><td>2,000</td></tr>
          <tr><td>취득</td><td>500</td></tr>
          <tr><td>기말</td><td>2,400</td></tr>
        </table>
        """,
        encoding="utf-8",
    )
    payload = {
        "source": str(source),
        "company": "Sample Co",
        "tolerance": 1,
        "summary": {"total": 2, "matched": 1, "unexplained_gap": 1},
        "results": [
            {
                "table_index": 0,
                "heading": "11. 유형자산 (1) 유형자산 변동내역입니다.",
                "status": "matched",
                "reason": "all columns foot within tolerance",
                "columns": [
                    {
                        "label": "합계",
                        "expected": 1000,
                        "actual": 1000,
                        "difference": 0,
                        "status": "matched",
                    }
                ],
            },
            {
                "table_index": 1,
                "heading": "12. 무형자산 (1) 무형자산 변동내역입니다.",
                "status": "unexplained_gap",
                "reason": "one or more columns do not foot",
                "columns": [
                    {
                        "label": "합계",
                        "expected": 2500,
                        "actual": 2400,
                        "difference": -100,
                        "status": "unexplained_gap",
                    }
                ],
            },
        ],
    }
    output = tmp_path / "sample_company_footing.xlsx"

    export_company_workbook(payload, output)

    workbook = load_workbook(output)
    assert workbook.sheetnames == [
        "Dashboard",
        "Note Summary",
        "Gap Review",
        "Note 11",
        "Note 12",
    ]
    assert workbook["Dashboard"]["A2"].value == "회사명"
    assert workbook["Dashboard"]["A4"].value == "전체 대사 항목"
    assert workbook["Dashboard"]["A6"].value == "미해소 차이"
    assert workbook["Dashboard"]["B2"].value == "Sample Co"
    assert workbook["Note Summary"]["A2"].value == "11"
    assert workbook["Note 11"]["E2"].value == "11"
    assert workbook["Note 12"]["K2"].value == -100
    assert workbook["Gap Review"]["E2"].value == "12"
