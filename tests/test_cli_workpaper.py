from openpyxl import load_workbook
from typer.testing import CliRunner

from dart_footing_reconciler.checks import CheckEvidence, CheckResult
from dart_footing_reconciler.cli import _run_workpaper_checks, app
from dart_footing_reconciler.document import FullReport, ReportBlock, ReportSection, ReportTable, SourceLocation
from dart_footing_reconciler.report_html import render_audit_reconciliation_html


def _section(section_id, title, kind, note_no, table):
    return ReportSection(
        section_id,
        title,
        kind,
        note_no,
        [ReportBlock("table", "", table, table.location)],
    )


def test_workpaper_checks_include_note_rollforward_assertions():
    report = FullReport(
        "sample.html",
        "Sample Co",
        [],
        [
            _section(
                "note:11",
                "유형자산",
                "note",
                "11",
                ReportTable(
                    0,
                    [
                        ["구분", "합계"],
                        ["기초장부금액", "100"],
                        ["취득", "10"],
                        ["기말장부금액", "110"],
                    ],
                    "유형자산의 변동내역 당기",
                    SourceLocation("note:11", 0, 0),
                ),
            )
        ],
    )

    checks = _run_workpaper_checks(report, None, tolerance=0)

    assert any(check.check_type == "note_rollforward_check" for check in checks)


def test_html_report_renders_note_assertion_section():
    report = FullReport("sample.html", "Sample Co", [], [])
    checks = [
        CheckResult(
            "note_assertion:11:table0:rollforward",
            "note_rollforward_check",
            "matched",
            "note",
            "11",
            "유형자산 증감표 검산",
            110,
            110,
            0,
            0,
            "기초와 변동내역이 기말 장부금액과 일치",
            [
                CheckEvidence("기초장부금액 합계", 100, "note:11/table:0/row:1/col:1"),
                CheckEvidence("기말장부금액 합계", 110, "note:11/table:0/row:3/col:1"),
            ],
        )
    ]

    html = render_audit_reconciliation_html(report, checks)

    assert "주석 내부/공식 검증" in html
    assert "유형자산 증감표 검산" in html
    assert "기초와 변동내역이 기말 장부금액과 일치" in html


def test_html_report_renders_asset_note_bridge_section():
    report = FullReport("sample.html", "Sample Co", [], [])
    checks = [
        CheckResult(
            "asset_note_bridge:intangible_assets.acquisitions_cashflow",
            "asset_note_bridge_check",
            "matched",
            "report",
            "14",
            "무형자산 취득 주석 연결 대사",
            80,
            80,
            0,
            0,
            "자산 주석과 관련 주석 금액이 현금흐름표 산식으로 연결됨",
            [
                CheckEvidence("cfs 무형자산의 취득", -80, "statement:cf/table:0/row:1/col:1"),
                CheckEvidence("note 14 취득", 100, "note:14/table:0/row:2/col:1"),
            ],
        )
    ]

    html = render_audit_reconciliation_html(report, checks)

    assert "주석 연결 대사" in html
    assert "무형자산 취득 주석 연결 대사" in html
    assert "자산 주석과 관련 주석 금액이 현금흐름표 산식으로 연결됨" in html


def test_cli_workpaper_excel_exports_note_sheets(tmp_path):
    source = tmp_path / "report.html"
    source.write_text(
        """
        <p>11. 유형자산</p>
        <p>유형자산 내용입니다.</p>
        <table><tr><th>구분</th><th>합계</th></tr><tr><td>기초</td><td>100</td></tr></table>
        """,
        encoding="utf-8",
    )
    output = tmp_path / "workpaper.xlsx"
    result = CliRunner().invoke(app, ["workpaper-excel", str(source), str(output), "--company", "Sample Co"])
    assert result.exit_code == 0
    wb = load_workbook(output)
    assert "Note 11" in wb.sheetnames


def test_cli_workpaper_excel_includes_required_check_types(tmp_path):
    current = tmp_path / "current.html"
    prior = tmp_path / "prior.html"
    current.write_text(
        """
        <p>재무상태표</p><table><tr><th>구분</th><th>당기</th></tr><tr><td>유형자산</td><td>1,000</td></tr><tr><td>매출채권</td><td>300</td></tr><tr><td>재고자산</td><td>200</td></tr></table>
        <p>손익계산서</p><table><tr><th>구분</th><th>당기</th></tr><tr><td>매출액</td><td>2,000</td></tr></table>
        <p>자본변동표</p><table><tr><th>구분</th><th>자본금</th><th>이익잉여금</th><th>합계</th></tr><tr><td>기초</td><td>100</td><td>700</td><td>800</td></tr><tr><td>배당</td><td>-</td><td>(100)</td><td>(100)</td></tr><tr><td>기말</td><td>100</td><td>600</td><td>700</td></tr></table>
        <p>현금흐름표</p><table><tr><th>구분</th><th>당기</th></tr><tr><td>유형자산의 취득</td><td>(500)</td></tr></table>
        <p>11. 유형자산</p><table><tr><th>구분</th><th>당기</th><th>전기</th><th>합계</th></tr><tr><td>기초</td><td>800</td><td>700</td><td>1,500</td></tr><tr><td>취득</td><td>500</td><td>400</td><td>900</td></tr><tr><td>장부금액</td><td>1,000</td><td>800</td><td>1,800</td></tr></table>
        <p>20. 고객과의 계약에서 생기는 수익</p><table><tr><th>구분</th><th>금액</th></tr><tr><td>매출액</td><td>2,000</td></tr></table>
        """,
        encoding="utf-8",
    )
    prior.write_text(
        """
        <p>10. 유형자산</p><table><tr><th>구분</th><th>당기</th></tr><tr><td>장부금액</td><td>800</td></tr></table>
        """,
        encoding="utf-8",
    )
    output = tmp_path / "workpaper.xlsx"
    result = CliRunner().invoke(
        app,
        ["workpaper-excel", str(current), str(output), "--company", "Sample Co", "--prior-html", str(prior)],
    )
    assert result.exit_code == 0
    wb = load_workbook(output)
    values = [cell.value for row in wb["Note 11"].iter_rows() for cell in row]
    assert "합계 검증 결과" in values
    assert "재무제표-주석 공식 계정 대사" in values
    assert "현금흐름표-주석 현금 변동 대사" in values
    assert "전기 공시 금액 대사" in values
    assert "전기말-당기초 대사" in values
    assert "재무제표-주석 대사" in values
    assert "현금흐름표-주석 직접 대사" in values


def test_cli_workpaper_html_exports_human_readable_reconciliation_report(tmp_path):
    current = tmp_path / "current.html"
    prior = tmp_path / "prior.html"
    current.write_text(
        """
        <p>재무상태표</p><table><tr><th>구분</th><th>당기</th></tr><tr><td>유형자산</td><td>1,000</td></tr><tr><td>매출채권</td><td>300</td></tr><tr><td>재고자산</td><td>200</td></tr></table>
        <p>손익계산서</p><table><tr><th>구분</th><th>당기</th></tr><tr><td>매출액</td><td>2,000</td></tr></table>
        <p>현금흐름표</p><table><tr><th>구분</th><th>당기</th></tr><tr><td>유형자산의 취득</td><td>(500)</td></tr></table>
        <p>11. 유형자산</p><table><tr><th>구분</th><th>당기</th><th>전기</th><th>합계</th></tr><tr><td>기초</td><td>800</td><td>700</td><td>1,500</td></tr><tr><td>취득</td><td>500</td><td>400</td><td>900</td></tr><tr><td>장부금액</td><td>1,000</td><td>800</td><td>1,800</td></tr></table>
        <p>20. 고객과의 계약에서 생기는 수익</p><table><tr><th>구분</th><th>금액</th></tr><tr><td>매출액</td><td>2,000</td></tr></table>
        """,
        encoding="utf-8",
    )
    prior.write_text(
        """
        <p>10. 유형자산</p><table><tr><th>구분</th><th>당기</th></tr><tr><td>장부금액</td><td>800</td></tr></table>
        """,
        encoding="utf-8",
    )
    output = tmp_path / "workpaper.html"

    result = CliRunner().invoke(
        app,
        ["workpaper-html", str(current), str(output), "--company", "Sample Co", "--prior-html", str(prior)],
    )

    assert result.exit_code == 0
    html = output.read_text(encoding="utf-8")
    assert "감사 대사 결과 보고서" in html
    assert 'href="#financial-position">재무상태표</a>' in html
    assert 'href="#income-statement">손익계산서</a>' in html
    assert 'href="#changes-in-equity">자본변동표</a>' in html
    assert 'href="#cash-flows">현금흐름표</a>' in html
    assert 'href="#notes">주석</a>' in html
    assert "전체 대사 항목" in html
    assert "미해소 차이" in html
    assert "검증 제외" in html
    assert "재무제표 원문" in html
    assert "근거 위치" in html
    assert 'class="leadsheet"' in html
    assert "원문 표" in html
    assert "선택 계정 주석" in html
    assert "note-drawer" in html
    assert "hover-note" in html
    assert "주석 원문 표" in html
    assert "전체 재무제표 계정 커버리지" in html
    assert "기말 잔액 직접 대사" not in html
    assert 'href="#balance"' not in html
    assert "보고서 순서 검증" in html
    assert "재무제표-주석 대사" in html
    assert "전기말-당기초 대사" in html
    assert "매출채권" in html
    assert "공식 계정 매핑 필요" in html
    assert "JSON" not in html
    assert "payload" not in html
    assert "check_type" not in html
    assert "account_key" not in html
    assert "note_rollforward_check" not in html
    assert "@media print" in html


def test_html_report_uses_audit_review_terms_instead_of_machine_statuses():
    checks = [
        CheckResult(
            "bs",
            "primary_balance_reconciliation",
            "unexplained_gap",
            "note",
            "11",
            "property_plant_equipment.balance",
            1_000,
            900,
            100,
            0,
            "financial statement line does not agree to note ending balance",
            [CheckEvidence("재무상태표 유형자산", 1_000, "statement:bs/table:0/row:1/col:1")],
        ),
        CheckResult(
            "cf",
            "cashflow_reconciliation",
            "unexplained_gap",
            "note",
            "11",
            "property_plant_equipment.acquisitions_cashflow",
            500,
            450,
            50,
            0,
            "cash flow statement amount does not agree to note movement",
            [CheckEvidence("현금흐름표 유형자산 취득", 500, "statement:cfs/table:0/row:1/col:1")],
        ),
        CheckResult(
            "total",
            "total_check",
            "parse_uncertain",
            "note",
            "11",
            "유형자산 total check",
            None,
            None,
            None,
            0,
            "no reliable total label found",
            [],
        ),
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", [], []), checks)

    assert "실질 차이 확인 필요" in html
    assert "차이내역 확인 필요" in html
    assert "합계 구조 확인 필요" not in html
    assert "unexplained_gap" not in html
    assert "파싱 불확실" not in html
    assert "unexplained_gap" not in html
    assert "parse_uncertain" not in html


def test_html_report_renders_validations_in_report_form_order():
    statements = [
        _section(
            "statement:cfs",
            "현금흐름표",
            "statement",
            "",
            ReportTable(0, [["구분", "당기"], ["유형자산의 취득", "(80)"]], "현금흐름표", SourceLocation("statement:cfs", 0, 0)),
        ),
        _section(
            "statement:bs",
            "재무상태표",
            "statement",
            "",
            ReportTable(0, [["구분", "당기"], ["유형자산", "100"]], "재무상태표", SourceLocation("statement:bs", 0, 0)),
        ),
        _section(
            "statement:pl",
            "손익계산서",
            "statement",
            "",
            ReportTable(0, [["구분", "당기"], ["매출액", "200"]], "손익계산서", SourceLocation("statement:pl", 0, 0)),
        ),
        _section(
            "statement:sce",
            "자본변동표",
            "statement",
            "",
            ReportTable(0, [["구분", "자본금", "합계"], ["기말", "50", "50"]], "자본변동표", SourceLocation("statement:sce", 0, 0)),
        ),
    ]
    notes = [
        _section(
            "note:11",
            "유형자산",
            "note",
            "11",
            ReportTable(0, [["구분", "당기", "합계"], ["장부금액", "100", "100"]], "유형자산", SourceLocation("note:11", 0, 0)),
        )
    ]
    checks = [
        CheckResult(
            "bs",
            "primary_balance_reconciliation",
            "matched",
            "report",
            "11",
            "property_plant_equipment.balance",
            100,
            100,
            0,
            0,
            "financial statement line agrees to note ending balance",
            [
                CheckEvidence("재무상태표 유형자산", 100, "statement:bs/table:0/row:1/col:1"),
                CheckEvidence("note 11 장부금액", 100, "note:11/table:0/row:1/col:1"),
            ],
        ),
        CheckResult(
            "cf",
            "cashflow_reconciliation",
            "matched",
            "report",
            "11",
            "property_plant_equipment.acquisitions_cashflow",
            80,
            80,
            0,
            0,
            "cash flow statement line agrees to note cash movement",
            [
                CheckEvidence("cfs 유형자산의 취득", -80, "statement:cfs/table:0/row:1/col:1"),
                CheckEvidence("note 11 취득", 80, "note:11/table:0/row:1/col:1"),
            ],
        ),
        CheckResult(
            "total",
            "total_check",
            "matched",
            "note",
            "11",
            "장부금액 row total",
            100,
            100,
            0,
            0,
            "row total agrees",
            [CheckEvidence("장부금액", 100, "note:11/table:0/row:1/col:2")],
        ),
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", statements, notes), checks)

    section_order = [
        html.index('id="financial-position"'),
        html.index('id="income-statement"'),
        html.index('id="changes-in-equity"'),
        html.index('id="cash-flows"'),
        html.index('id="notes"'),
    ]
    assert section_order == sorted(section_order)
    financial_position = html[section_order[0] : section_order[1]]
    income_statement = html[section_order[1] : section_order[2]]
    equity_statement = html[section_order[2] : section_order[3]]
    cash_flows = html[section_order[3] : section_order[4]]
    notes_section = html[section_order[4] : html.index('data-view-panel="review"')]
    assert "재무제표-주석 대사" in financial_position
    assert "유형자산" in financial_position
    assert "자동 검증 미수행" in income_statement
    assert "자동화 보완 필요" in equity_statement
    assert "CF↔주석 대사" in cash_flows
    assert "유형자산 취득" in cash_flows
    assert "합계 검증" in notes_section
    assert "장부금액 row total" not in html


def test_html_report_first_viewport_is_worksheet_cover_with_tickmark_legend():
    checks = [
        CheckResult(
            "bs",
            "primary_balance_reconciliation",
            "unexplained_gap",
            "note",
            "11",
            "property_plant_equipment.balance",
            1_000,
            900,
            100,
            0,
            "financial statement line does not agree to note ending balance",
            [CheckEvidence("재무상태표 유형자산", 1_000, "statement:bs/table:0/row:1/col:1")],
        ),
        CheckResult(
            "total",
            "total_check",
            "parse_uncertain",
            "note",
            "11",
            "유형자산 total check",
            None,
            None,
            None,
            0,
            "no reliable total label found",
            [],
        )
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", [], []), checks)
    cover = html.index('<header class="report-header worksheet-cover"')
    financial_position = html.index('id="financial-position"')
    first_viewport = html[cover:financial_position]

    # 보고서 표지: 결론 + 작성자/검토자 signoff
    assert "감사 대사 결과 보고서" in first_viewport
    assert '<dl class="signoff"' in first_viewport
    assert "작성자" in first_viewport
    assert "검토자" in first_viewport
    # 검산 표기 범례 (tickmark legend)
    assert "검산 표기" in first_viewport
    assert "재무제표·주석 일치" in first_viewport
    assert "현금흐름 대사" in first_viewport
    # working / review 탭 분리
    assert 'data-view-tab="working"' in first_viewport
    assert 'data-view-tab="review"' in first_viewport
    assert "감사 대사 결과" in first_viewport
    assert "리뷰 요약" in first_viewport
    # 제품 대시보드 chrome 제거 확인
    assert '<section class="section-brief"' not in html
    assert "왜 중요한가" not in html
    # working/review 패널 분리 + 리뷰 큐는 review 패널로 이동
    assert 'data-view-panel="working"' in html
    assert 'data-view-panel="review"' in html
    assert html.index('data-view-panel="review"') > html.index('id="financial-position"')


def test_html_status_badges_use_point_signal_not_colored_fills():
    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", [], []), [])

    assert ".status::before" in html
    assert ".status-ok { border-color: var(--ok); color: var(--ok); }" in html
    assert ".status-risk { border-color: var(--risk); color: var(--risk); }" in html
    assert ".status-ok { background: var(--ok-soft); color: var(--ok); }" not in html
    assert ".status-risk { background: var(--risk-soft); color: var(--risk); }" not in html


def test_html_report_uses_design_kit_spacing_and_section_surfaces():
    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", [], []), [])

    assert "letter-spacing: 0;" in html
    assert "letter-spacing: -0.02em;" not in html
    assert ".section-brief, .report-section { margin-top: 26px; padding-top: 22px; border-top: 1px solid var(--line); }" in html
    assert ".section-brief, .report-section { background: var(--surface); border: 1px solid var(--line);" not in html
    assert ".report-header { display: flex; justify-content: space-between; gap: 20px; align-items: flex-start; padding: 8px 0 18px; border-bottom: 1px solid var(--line); }" in html


def test_html_report_translates_internal_evidence_terms_for_reviewers():
    checks = [
        CheckResult(
            "cf",
            "cashflow_reconciliation",
            "explainable_gap",
            "note",
            "11",
            "property_plant_equipment.acquisitions_cashflow",
            800,
            700,
            -100,
            0,
            "주석 취득 1,000 - 비현금거래-미지급금 증가 300 = 700; 현금흐름표 유형자산의 취득 800; 차이 (100); 현금흐름표 금액과 직접 대사되지 않음",
            [
                CheckEvidence("cfs 유형자산의 취득", 800, "statement:현금흐름표/table:0/row:1/col:1"),
                CheckEvidence("note 11 취득", 1_000, "note:11/table:0/row:2/col:1"),
            ],
        ),
        CheckResult(
            "note-note",
            "note_note_match",
            "unexplained_gap",
            "note",
            "14",
            "amortization_expense note to note match",
            100,
            90,
            10,
            0,
            "financial statement amount does not agree to note amount",
            [CheckEvidence("note 14 상각비", 100, "note:14/table:2/row:3/col:1")],
        ),
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", [], []), checks)

    assert "현금흐름표 유형자산의 취득" in html
    assert "주석 11 취득" in html
    assert "주석 14 · 표 2 · 행 3 · 열 1" in html
    assert "무형자산상각비 주석 간 대사" in html
    assert "cfs 유형자산의 취득" not in html
    assert "note 11 취득" not in html
    assert "note:14/table:2/row:3/col:1" not in html
    assert "amortization_expense note to note match" not in html


def test_html_cashflow_formula_reason_is_rendered_as_audit_table():
    checks = [
        CheckResult(
            "cf",
            "cashflow_reconciliation",
            "explainable_gap",
            "note",
            "11",
            "property_plant_equipment.acquisitions_cashflow",
            800,
            700,
            -100,
            0,
            "주석 취득 1,000 - 비현금거래-미지급금 증가 300 = 700; 현금흐름표 유형자산의 취득 800; 차이 (100); 현금흐름표 금액과 직접 대사되지 않음",
            [
                CheckEvidence("cfs 유형자산의 취득", 700, "statement:cfs/table:0/row:1/col:1"),
                CheckEvidence("note 11 취득", 1_000, "note:11/table:0/row:1/col:1"),
            ],
        )
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", [], []), checks)
    review_queue = html[html.index('<section class="report-section review-queue"') :]
    review_queue = review_queue[: review_queue.index("</section>")]

    assert '<table class="formula-table">' in review_queue
    assert 'class="review-action-formula"' in review_queue
    assert "<th>구분</th>" in review_queue
    assert "<th>내용</th>" in review_queue
    assert "<td>산식</td>" in review_queue
    assert "<td>대사 대상</td>" in review_queue
    assert "formula-row" not in review_queue


def test_html_report_shows_note_total_check_section_with_subtotal_differences():
    notes = [
        _section(
            "note:11",
            "유형자산",
            "note",
            "11",
            ReportTable(
                0,
                [["구분", "토지", "건물", "합계"], ["기초", "100", "200", "301"], ["소계", "200", "300", "500"]],
                "11. 유형자산",
                SourceLocation("note:11", 0, 0),
            ),
        )
    ]
    checks = [
        CheckResult(
            "total:11:table0:row1",
            "total_check",
            "unexplained_gap",
            "note",
            "11",
            "기초 row total",
            300,
            301,
            1,
            0,
            "row total does not agree",
            [CheckEvidence("기초", 301, "note:11/table:0/row:1/col:3")],
        ),
        CheckResult(
            "total:11:table0:row2",
            "total_check",
            "matched",
            "note",
            "11",
            "소계 row total",
            500,
            500,
            0,
            0,
            "row total agrees",
            [CheckEvidence("소계", 500, "note:11/table:0/row:2/col:3")],
        ),
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", [], notes), checks)

    assert 'id="notes"' in html
    assert "주석 원문 및 검증" in html
    assert "검증 가능 항목 2개 중 일치 1개, 차이 1개" in html
    assert "합계 차이 확인 필요" in html
    assert "일치 1 · 차이 1" in html
    assert "구성항목 합계" in html
    assert "표시 소계/합계" in html
    assert "주석 11. 유형자산" in html
    assert "total-issue-cell" in html
    assert "행 소계/합계 · 구성항목 합계 300 · 표시 소계/합계 301 · 차이 1" in html


def test_html_report_renders_note_workspace_with_text_and_comparison_panels():
    statements = [
        _section(
            "statement:bs",
            "재무상태표",
            "statement",
            "",
            ReportTable(0, [["구분", "당기"], ["유형자산", "1,000"]], "재무상태표", SourceLocation("statement:bs", 0, 0)),
        ),
        _section(
            "statement:pl",
            "포괄손익계산서",
            "statement",
            "",
            ReportTable(1, [["구분", "당기"], ["감가상각비", "100"]], "포괄손익계산서", SourceLocation("statement:pl", 0, 1)),
        ),
        _section(
            "statement:sce",
            "자본변동표",
            "statement",
            "",
            ReportTable(2, [["구분", "이익잉여금"], ["배당", "(50)"]], "자본변동표", SourceLocation("statement:sce", 0, 2)),
        ),
        _section(
            "statement:cfs",
            "현금흐름표",
            "statement",
            "",
            ReportTable(3, [["구분", "당기"], ["유형자산의 취득", "(200)"]], "현금흐름표", SourceLocation("statement:cfs", 0, 3)),
        ),
    ]
    note_table = ReportTable(
        4,
        [["구분", "당기"], ["장부금액", "1,000"], ["취득", "200"], ["감가상각비", "100"]],
        "유형자산 변동내역",
        SourceLocation("note:11", 1, 4),
    )
    other_note_table = ReportTable(
        5,
        [["구분", "당기"], ["감가상각비", "100"]],
        "비용의 성격별 분류",
        SourceLocation("note:20", 0, 5),
    )
    notes = [
        ReportSection(
            "note:11",
            "유형자산",
            "note",
            "11",
            [
                ReportBlock("text", "회사의 유형자산은 취득원가에서 감가상각누계액을 차감하여 표시합니다.", None, SourceLocation("note:11", 0)),
                ReportBlock("table", "", note_table, note_table.location),
            ],
        ),
        _section("note:20", "비용의 성격별 분류", "note", "20", other_note_table),
    ]
    checks = [
        CheckResult(
            "bs-note",
            "primary_balance_reconciliation",
            "matched",
            "report",
            "11",
            "property_plant_equipment.balance",
            1_000,
            1_000,
            0,
            0,
            "financial statement line agrees to note ending balance",
            [
                CheckEvidence("재무상태표 유형자산", 1_000, "statement:bs/table:0/row:1/col:1"),
                CheckEvidence("note 11 장부금액", 1_000, "note:11/table:4/row:1/col:1"),
            ],
        ),
        CheckResult(
            "pl-note",
            "expense_allocation",
            "matched",
            "report",
            "11",
            "property_plant_equipment.depreciation_expense_allocation",
            100,
            100,
            0,
            0,
            "financial statement amount agrees to note amount",
            [
                CheckEvidence("포괄손익계산서 감가상각비", 100, "statement:pl/table:1/row:1/col:1"),
                CheckEvidence("note 11 감가상각비", 100, "note:11/table:4/row:3/col:1"),
            ],
        ),
        CheckResult(
            "sce-note",
            "note_note_match",
            "matched",
            "note",
            "11",
            "dividends note to equity match",
            50,
            50,
            0,
            0,
            "financial statement amount agrees to note amount",
            [
                CheckEvidence("자본변동표 배당", -50, "statement:sce/table:2/row:1/col:1"),
                CheckEvidence("note 11 배당", 50, "note:11/table:4/row:2/col:1"),
            ],
        ),
        CheckResult(
            "cf-note",
            "cashflow_reconciliation",
            "matched",
            "report",
            "11",
            "property_plant_equipment.acquisitions_cashflow",
            200,
            200,
            0,
            0,
            "cash flow statement line agrees to note cash movement",
            [
                CheckEvidence("cfs 유형자산의 취득", -200, "statement:cfs/table:3/row:1/col:1"),
                CheckEvidence("note 11 취득", 200, "note:11/table:4/row:2/col:1"),
            ],
        ),
        CheckResult(
            "note-note",
            "note_note_match",
            "matched",
            "note",
            "11",
            "depreciation expense note to note match",
            100,
            100,
            0,
            0,
            "financial statement amount agrees to note amount",
            [
                CheckEvidence("note 11 감가상각비", 100, "note:11/table:4/row:3/col:1"),
                CheckEvidence("note 20 감가상각비", 100, "note:20/table:5/row:1/col:1"),
            ],
        ),
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", statements, notes), checks)
    notes_section = html[html.index('id="notes"') : html.index('data-view-panel="review"')]

    assert 'class="note-workspace"' in notes_section
    assert 'data-note-tab="note-11"' in notes_section
    assert 'data-note-panel="note-11"' in notes_section
    assert "회사의 유형자산은 취득원가에서 감가상각누계액을 차감하여 표시합니다." in notes_section
    assert "재무상태표 연결" in notes_section
    assert "손익계산서 연결" in notes_section
    assert "자본변동표 연결" in notes_section
    assert "현금흐름표 연결" in notes_section
    assert "다른 주석 대사" in notes_section
    assert "statement-source-preview" in notes_section
    assert "유형자산의 취득" in notes_section
    assert "비용의 성격별 분류" in notes_section


def test_html_cashflow_review_queue_shows_formula_result_not_instructions():
    checks = [
        CheckResult(
            "cf",
            "cashflow_reconciliation",
            "explainable_gap",
            "note",
            "11",
            "property_plant_equipment.acquisitions_cashflow",
            800,
            700,
            -100,
            0,
            "주석 취득 1,000 - 비현금거래-미지급금 증가 300 = 700; 현금흐름표 유형자산의 취득 800; 차이 (100); 현금흐름표 금액과 직접 대사되지 않음",
            [
                CheckEvidence("cfs 유형자산의 취득", 800, "statement:cfs/table:0/row:1/col:1"),
                CheckEvidence("note 11 취득", 1_000, "note:11/table:0/row:1/col:1"),
                CheckEvidence("note 11 비현금거래-미지급금", 300, "note:11/table:0/row:2/col:1"),
            ],
        )
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", [], []), checks)
    review_queue = html[html.index('<section class="report-section review-queue"') :]
    review_queue = review_queue[: review_queue.index("</section>")]

    assert "대사 결과" in review_queue
    assert "주석 취득 1,000 - 비현금거래-미지급금 증가 300 = 700" in review_queue
    assert "현금흐름표 유형자산의 취득 800" in review_queue
    assert "산식화하세요" not in review_queue
    assert "보완하세요" not in review_queue
    assert "확인하세요" not in review_queue


def test_cli_workpaper_html_hover_uses_current_period_and_raw_note_table(tmp_path):
    current = tmp_path / "current.html"
    current.write_text(
        """
        <p>재무상태표</p>
        <table>
          <tr><th></th><th>제 114 기</th><th>제 113 기</th><th>제 112 기</th></tr>
          <tr><td>유형자산</td><td>3,000</td><td>2,000</td><td>1,000</td></tr>
        </table>
        <p>11. 유형자산</p>
        <table>
          <tr><th>구분</th><th>합계</th></tr>
          <tr><td>기말 순장부금액</td><td>3,000</td></tr>
        </table>
        """,
        encoding="utf-8",
    )
    output = tmp_path / "workpaper.html"

    result = CliRunner().invoke(
        app,
        ["workpaper-html", str(current), str(output), "--company", "Sample Co"],
    )

    assert result.exit_code == 0
    html = output.read_text(encoding="utf-8")
    assert "<b>본문 금액</b>3,000" in html
    assert "<b>본문 금액</b>1,000" not in html
    assert "주석 원문 표" in html
    assert "기말 순장부금액" in html


def test_html_hover_excludes_policy_risk_note_and_pairs_current_prior_amounts():
    statement = _section(
        "statement:bs",
        "재무상태표",
        "statement",
        "",
        ReportTable(
            0,
            [["구분", "당기"], ["당기손익-공정가치측정금융자산", "1,220,780,000"]],
            "재무상태표",
            SourceLocation("statement:bs", 0, 0),
            row_acodes=[
                ["||||", "||||"],
                ["||||", "ifrs-full_CurrentFinancialAssetsAtFairValueThroughProfitOrLoss|CFY|0|KRW|"],
            ],
        ),
    )
    notes = [
        _section(
            "note:risk",
            "중요한 회계정책",
            "note",
            "2",
            ReportTable(
                1,
                [["구분", "최대노출금액"], ["당기손익-공정가치측정금융자산", "1,220,780"]],
                "2. 중요한 회계정책 (2) 신용위험",
                SourceLocation("note:risk", 0, 1),
                unit_multiplier=1000,
            ),
        ),
        _section(
            "note:fair-value",
            "금융상품",
            "note",
            "7",
            ReportTable(
                2,
                [["구분", "공정가치"], ["당기손익-공정가치측정금융자산", "1,220,780"]],
                "7. 가치평가기법 및 투입변수",
                SourceLocation("note:fair-value", 0, 2),
                unit_multiplier=1000,
            ),
        ),
        _section(
            "note:financial-assets",
            "금융상품",
            "note",
            "7",
            ReportTable(
                3,
                [["구분", "당기"], ["당기손익-공정가치측정금융자산", "1,220,780"]],
                "7. 금융자산의 범주별 장부금액",
                SourceLocation("note:financial-assets", 0, 3),
                unit_multiplier=1000,
            ),
        ),
        _section(
            "note:financial-assets",
            "금융상품",
            "note",
            "7",
            ReportTable(
                4,
                [["구분", "전기"], ["당기손익-공정가치측정금융자산", "900,000"]],
                "7. 금융자산의 범주별 장부금액",
                SourceLocation("note:financial-assets", 1, 4),
                unit_multiplier=1000,
            ),
        ),
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", [statement], notes), [])

    assert "매칭 주석" in html
    assert "금융자산의 범주별 장부금액" in html
    assert "1,220,780,000" in html
    notes_start = html.index('id="notes"')
    note_match_html = html[html.index("매칭 주석") : notes_start]
    assert "900,000" in note_match_html
    assert "900,000,000" not in note_match_html
    assert note_match_html.index("금융자산의 범주별 장부금액") < note_match_html.index("가치평가기법")
    assert "최대노출금액" in html


def test_html_statement_match_uses_each_statement_row_amount_and_scope():
    statement = _section(
        "statement:bs",
        "재무상태표",
        "statement",
        "",
        ReportTable(
            0,
            [["구분", "당기"], ["현금및현금성자산", "290,135,224,989"], ["현금및현금성자산", "79,940,581,000"]],
            "재무상태표",
            SourceLocation("statement:bs", 0, 0),
            row_acodes=[
                ["||||", "||||"],
                ["ifrs-full_CashAndCashEquivalents|CFY|ConsolidatedMember|KRW|", "||||"],
                ["ifrs-full_CashAndCashEquivalents|CFY|SeparateMember|KRW|", "||||"],
            ],
        ),
    )
    notes = [
        _section(
            "note:7",
            "범주별 금융상품 (연결)",
            "note",
            "7",
            ReportTable(
                1,
                [["구분", "당기", "전기"], ["현금및현금성자산", "290,135,225", "582,914,149"]],
                "7. 범주별 금융상품 (연결) 금융자산의 범주별 장부금액",
                SourceLocation("note:7", 0, 1),
                unit_multiplier=1000,
            ),
        ),
        _section(
            "note:6",
            "범주별 금융상품",
            "note",
            "6",
            ReportTable(
                2,
                [["구분", "당기", "전기"], ["현금및현금성자산", "79,940,581", "338,843,218"]],
                "6. 범주별 금융상품 금융자산의 범주별 장부금액",
                SourceLocation("note:6", 0, 2),
                unit_multiplier=1000,
            ),
        ),
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", [statement], notes), [])

    first_panel = html[html.index("<b>본문 금액</b>290,135,224,989") :]
    first_panel = first_panel[: first_panel.index("</button>")]
    second_panel = html[html.index("<b>본문 금액</b>79,940,581,000") :]
    second_panel = second_panel[: second_panel.index("</button>")]
    assert "290,135,225" in first_panel
    assert "290,135,225,000" not in first_panel
    assert "79,940,581,000" not in first_panel
    assert "7. 범주별 금융상품 (연결)" in first_panel
    assert "6. 범주별 금융상품 금융자산" not in first_panel
    assert "79,940,581" in second_panel
    assert "290,135,225,000" not in second_panel


def test_html_income_statement_source_table_preserves_report_rows():
    statement = _section(
        "statement:pl",
        "손익계산서",
        "statement",
        "",
        ReportTable(
            0,
            [
                ["구분", "당기"],
                ["손익/기타포괄손익 개별작성", ""],
                ["세후기타포괄손익", ""],
                ["매출액", "1,000"],
                ["매출원가", "(700)"],
                ["당기순이익", "100"],
            ],
            "손익계산서",
            SourceLocation("statement:pl", 0, 0),
        ),
    )

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", [statement], []), [])
    block = html[html.index("<h3>손익계산서</h3>") :]
    block = block[: block.index("</article>")]

    assert "매출액" in block
    assert "매출원가" in block
    assert block.index("매출액") < block.index("매출원가")
    assert "손익/기타포괄손익 개별작성" not in block
    assert "세후기타포괄손익" not in block


def test_html_statement_match_falls_back_to_note_label_and_unit_rounding():
    statement = _section(
        "statement:bs",
        "재무상태표",
        "statement",
        "",
        ReportTable(
            0,
            [["구분", "당기"], ["기타금융자산", "487,979,917,360"]],
            "재무상태표",
            SourceLocation("statement:bs", 0, 0),
        ),
    )
    notes = [
        _section(
            "note:4",
            "범주별 금융상품",
            "note",
            "4",
            ReportTable(
                1,
                [
                    ["구분", "상각후원가", "당기손익-공정가치 측정 금융자산", "합계"],
                    ["기타금융자산(*1)", "0", "487,980", "487,980"],
                ],
                "4. 범주별 금융상품 금융자산의 범주별 공시 당기 (단위 : 백만원)",
                SourceLocation("note:4", 0, 1),
                unit_multiplier=1_000_000,
            ),
        )
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", [statement], notes), [])

    panel = html[html.index("<b>본문 금액</b>487,979,917,360") :]
    panel = panel[: panel.index("</button>")]
    assert "<b>연결 주석</b>범주별 금융상품 금융자산의 범주별 공시 당기 (단위 : 백만원) · 기타금융자산(*1)" in panel
    assert "487,980" in panel
    assert "487,980,000,000" not in panel
    assert "주석 원문 표" in panel


def test_html_note_candidate_table_uses_actual_amounts_not_raw_table_placeholder():
    statement = _section(
        "statement:bs",
        "재무상태표",
        "statement",
        "",
        ReportTable(
            0,
            [["구분", "당기"], ["기타유동금융자산", "14,826,190,589"]],
            "재무상태표",
            SourceLocation("statement:bs", 0, 0),
        ),
    )
    notes = [
        _section(
            "note:financial-risk",
            "재무위험관리",
            "note",
            "29",
            ReportTable(
                1,
                [
                    ["구분", "당기", "전기"],
                    ["기타유동금융자산", "29,607,779,689", "24,428,120,463"],
                    ["범주별 금융상품", "213,111,100,820", ""],
                ],
                "재무위험관리 (연결) 금융자산의 범주별 공시 당기 (단위 : 원)",
                SourceLocation("note:financial-risk", 0, 1),
            ),
        )
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", [statement], notes), [])
    panel = html[html.index("<b>본문 금액</b>14,826,190,589") :]
    panel = panel[: panel.index("</button>")]

    assert "29,607,779,689" in panel
    assert "24,428,120,463" in panel
    assert '<span class="note-match-cell note-match-num">원문 표</span>' not in panel


def test_html_equity_statement_header_uses_capital_layer_accounts():
    statement = _section(
        "statement:ce",
        "자본변동표",
        "statement",
        "",
        ReportTable(
            0,
            [
                ["구분", "자본", "자본", "자본", "자본"],
                ["구분", "자본금", "자본잉여금", "기타포괄손익누계액", "이익잉여금"],
                ["2024.01.01 (기초자본)", "100", "200", "300", "400"],
            ],
            "자본변동표",
            SourceLocation("statement:ce", 0, 0),
        ),
    )

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", [statement], []), [])
    block = html[html.index("<h3>자본변동표</h3>") :]
    block = block[: block.index("</article>")]

    assert "<th>자본금</th>" in block
    assert "<th>자본잉여금</th>" in block
    assert "<th>기타포괄손익누계액</th>" in block
    assert "<th>이익잉여금</th>" in block
    assert block.count("<th>자본</th>") == 0


def test_html_report_displays_zero_amounts_as_dash_in_raw_tables_and_amount_cells():
    statement = _section(
        "statement:cfs",
        "현금흐름표",
        "statement",
        "",
        ReportTable(
            0,
            [["구분", "당기"], ["무형자산의 취득", "(551,507,770)"]],
            "현금흐름표",
            SourceLocation("statement:cfs", 0, 0),
        ),
    )
    notes = [
        _section(
            "note:intangibles",
            "무형자산",
            "note",
            "14",
            ReportTable(
                1,
                [["구분", "저작권", "회원권"], ["취득", "0", "551,507,770"], ["대체", "", "0"]],
                "무형자산 변동내역",
                SourceLocation("note:intangibles", 0, 1),
            ),
        )
    ]
    checks = [
        CheckResult(
            "cf",
            "cashflow_reconciliation",
            "matched",
            "note",
            "14",
            "intangible_assets.acquisitions_cashflow",
            551_507_770,
            551_507_770,
            0,
            0,
            "주석 취득 551,507,770 = 551,507,770; 현금흐름표 무형자산의 취득 551,507,770; 차이 0; 현금흐름표 금액과 직접 대사됨",
            [
                CheckEvidence("cfs 무형자산의 취득", -551_507_770, "statement:cfs/table:0/row:1/col:1"),
                CheckEvidence("note 14 취득", 551_507_770, "note:14/table:1/row:1/col:2"),
            ],
        )
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", [statement], notes), checks)

    assert "<td>-</td>" in html
    assert "<td>0</td>" not in html
    assert "차이</td><td>-</td>" in html


def test_html_cashflow_hover_judgment_renders_formula_as_rows():
    statement = _section(
        "statement:cfs",
        "현금흐름표",
        "statement",
        "",
        ReportTable(
            0,
            [["구분", "당기"], ["무형자산의 취득", "(551,507,770)"]],
            "현금흐름표",
            SourceLocation("statement:cfs", 0, 0),
        ),
    )
    checks = [
        CheckResult(
            "cf",
            "cashflow_reconciliation",
            "matched",
            "note",
            "14",
            "intangible_assets.acquisitions_cashflow",
            551_507_770,
            551_507_770,
            0,
            0,
            "주석 취득 551,507,770 = 551,507,770; 현금흐름표 무형자산의 취득 551,507,770; 차이 0; 현금흐름표 금액과 직접 대사됨",
            [
                CheckEvidence("cfs 무형자산의 취득", -551_507_770, "statement:cfs/table:0/row:1/col:1"),
                CheckEvidence("note 14 취득", 551_507_770, "note:14/table:1/row:1/col:2"),
            ],
        )
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", [statement], []), checks)
    panel = html[html.index("<b>본문 금액</b>(551,507,770)") :]
    panel = panel[: panel.index("</button>")]

    assert 'class="judgment-block"' in panel
    assert '<table class="formula-table">' in panel
    assert "주석 취득 551,507,770 = 551,507,770; 현금흐름표" not in panel


def test_html_income_statement_links_non_operating_and_finance_notes():
    statement = _section(
        "statement:pl",
        "손익계산서",
        "statement",
        "",
        ReportTable(
            0,
            [
                ["구분", "당기"],
                ["영업외수익", "4,004,885,291"],
                ["영업외비용", "1,961,843,947"],
                ["금융수익", "74,447,542,617"],
                ["금융비용", "32,234,009,126"],
            ],
            "손익계산서",
            SourceLocation("statement:pl", 0, 0),
        ),
    )
    notes = [
        _section(
            "note:other",
            "영업외손익",
            "note",
            "27",
            ReportTable(
                1,
                [["구분", "당기"], ["영업외수익", "4,004,885,291"], ["영업외비용", "1,961,843,947"]],
                "영업외수익 및 영업외비용 상세내역",
                SourceLocation("note:other", 0, 1),
            ),
        ),
        _section(
            "note:finance",
            "금융손익",
            "note",
            "28",
            ReportTable(
                2,
                [["구분", "당기"], ["금융수익", "74,447,542,617"], ["금융비용", "32,234,009,126"]],
                "금융수익 및 금융비용 상세내역",
                SourceLocation("note:finance", 0, 2),
            ),
        ),
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", [statement], notes), [])

    assert "영업외수익 및 영업외비용 상세내역 · 영업외비용" in html
    assert "금융수익 및 금융비용 상세내역 · 금융수익" in html
    assert "<b>연결 주석</b>미연결" not in html


def test_html_report_separates_consolidated_and_separate_statement_scopes():
    statements = [
        _section(
            "statement:bs",
            "재무상태표",
            "statement",
            "",
            ReportTable(
                0,
                [["구분", "당기"], ["현금및현금성자산", "290,000"]],
                "재무상태표",
                SourceLocation("statement:bs", 0, 0),
            ),
        ),
        _section(
            "statement:cf",
            "현금흐름표",
            "statement",
            "",
            ReportTable(
                1,
                [["구분", "당기"], ["현금및현금성자산의증가", "10"]],
                "현금흐름표",
                SourceLocation("statement:cf", 0, 1),
            ),
        ),
        _section(
            "statement:bs",
            "재무상태표",
            "statement",
            "",
            ReportTable(
                2,
                [["구분", "당기"], ["현금및현금성자산", "120,000"]],
                "재무상태표",
                SourceLocation("statement:bs", 1, 2),
            ),
        ),
        _section(
            "statement:cf",
            "현금흐름표",
            "statement",
            "",
            ReportTable(
                3,
                [["구분", "당기"], ["현금및현금성자산의증가", "5"]],
                "현금흐름표",
                SourceLocation("statement:cf", 1, 3),
            ),
        ),
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", statements, []), [])

    assert 'data-scope-tab="consolidated"' in html
    assert 'data-scope-tab="separate"' in html
    assert '<article class="statement-block report-frame-block" data-report-scope="consolidated">' in html
    assert '<article class="statement-block report-frame-block" data-report-scope="separate">' in html
    assert "<h3>연결 · 재무상태표</h3>" in html
    assert "<h3>별도 · 재무상태표</h3>" in html


def test_cli_workpaper_html_hover_shows_generic_fsc_account_without_note_candidate(tmp_path):
    current = tmp_path / "current.html"
    current.write_text(
        """
        <p>재무상태표</p>
        <table>
          <tr><th>구분</th><th>당기</th></tr>
          <tr><td acode="ifrs-full_CashAndCashEquivalents|CFY|0|KRW|">현금및현금성자산</td><td>290,135</td></tr>
        </table>
        """,
        encoding="utf-8",
    )
    output = tmp_path / "workpaper.html"

    result = CliRunner().invoke(
        app,
        ["workpaper-html", str(current), str(output), "--company", "Sample Co"],
    )

    assert result.exit_code == 0
    html = output.read_text(encoding="utf-8")
    assert "<b>공식 계정단</b>현금및현금성자산" in html
    assert "<b>공식 계정단</b>분류 필요" not in html
    assert "공식 계정 확인" in html
    assert "재무제표 계정명 기준 공식 계정단 확인됨" in html


def test_cli_workpaper_html_generic_fsc_uses_numeric_note_schema_not_risk_policy_table(tmp_path):
    current = tmp_path / "current.html"
    current.write_text(
        """
        <p>재무상태표</p>
        <table>
          <tr><th>구분</th><th>당기</th></tr>
          <tr><td acode="ifrs-full_CurrentFinancialAssetsAtFairValueThroughProfitOrLoss|CFY|0|KRW|">당기손익-공정가치측정금융자산</td><td>1,220,780</td></tr>
        </table>
        <p>2. 중요한 회계정책</p>
        <p>(2) 신용위험</p>
        <table>
          <tr><th>구분</th><th>최대노출금액</th></tr>
          <tr><td>당기손익-공정가치측정금융자산</td><td>1,220,780</td></tr>
        </table>
        <p>7. 금융상품</p>
        <table>
          <tr><th>구분</th><th>상각후원가</th><th>당기손익인식금융자산</th><th>합계</th></tr>
          <tr><td>당기손익-공정가치측정금융자산</td><td>0</td><td>1,220,780</td><td>49,153,191</td></tr>
        </table>
        """,
        encoding="utf-8",
    )
    output = tmp_path / "workpaper.html"

    result = CliRunner().invoke(
        app,
        ["workpaper-html", str(current), str(output), "--company", "Sample Co"],
    )

    assert result.exit_code == 0
    html = output.read_text(encoding="utf-8")
    assert "매칭 주석" in html
    assert '<span class="note-match-head">all</span>' in html


def test_cashflow_map_infers_periods_and_cashflow_note_without_company_specific_form():
    statements = [
        _section(
            "statement:cfs",
            "현금흐름표",
            "statement",
            "",
            ReportTable(
                0,
                [
                        ["구분", "제 42 기", "제 41 기"],
                        ["투자활동으로 인한 현금흐름", "", ""],
                        ["기타유동금융자산의 취득", "(2,000,000,000)", "(999,000,000,000)"],
                        ["재무활동으로 인한 현금흐름", "", ""],
                        ["배당금지급", "(1,500,000,000)", "(30,000,000,000)"],
                ],
                "현금흐름표",
                SourceLocation("statement:cfs", 0, 0),
            ),
        )
    ]
    notes = [
        _section(
            "note:finance-assets",
            "금융상품",
            "note",
            "8",
            ReportTable(
                2,
                [
                    ["구분", "제 42 기"],
                    ["기타유동금융자산의 취득", "2,000"],
                ],
                "기타금융자산 변동내역 (단위 : 백만원)",
                SourceLocation("note:finance-assets", 0, 2),
                unit_multiplier=1_000_000,
            ),
        ),
        _section(
            "note:dividend",
            "이익잉여금",
            "note",
            "21",
            ReportTable(
                3,
                [
                    ["구분", "제 42 기"],
                    ["배당금지급", "1,500"],
                ],
                "이익잉여금 배당 내역 (단위 : 백만원)",
                SourceLocation("note:dividend", 0, 3),
                unit_multiplier=1_000_000,
            ),
        ),
        _section(
            "note:prior-borrowing",
            "차입금",
            "note",
            "22",
            ReportTable(
                4,
                [
                    ["구분", "제 41 기"],
                    ["기타유동금융자산의 취득", "2,000"],
                ],
                "차입금 전기 (단위 : 백만원)",
                SourceLocation("note:prior-borrowing", 0, 4),
                unit_multiplier=1_000_000,
            ),
        ),
        _section(
            "note:cashflow",
            "현금흐름",
            "note",
            "9",
            ReportTable(
                1,
                [
                    ["구분", "제 42 기", "제 41 기"],
                    ["비현금항목의 조정 법인세비용", "2,000,000,000", "10"],
                ],
                "영업에서 창출된 현금흐름",
                SourceLocation("note:cashflow", 0, 1),
            ),
        )
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", statements, notes), [])
    start = html.index('<section class="report-section report-frame-section" id="cash-flows"')
    cashflow_map = html[start : html.index('<section class="report-section report-frame-section" id="notes"', start)]

    assert "기타유동금융자산의 취득" in cashflow_map
    assert "(2,000,000,000)" in cashflow_map
    validation_panel = cashflow_map[cashflow_map.index("CF↔주석 대사") :]
    assert "(999,000,000,000)" not in validation_panel
    assert "법인세비용" in cashflow_map
    assert "2,000,000,000" in cashflow_map
    assert "손익계산서 법인세비용 + 법인세비용 주석" in cashflow_map
    assert "현금흐름표" in cashflow_map
    assert "현금흐름 대사 범위" in cashflow_map
    assert "자본변동표 + 자본/배당 주석" in cashflow_map
    assert "주석에서 확인된 금액" in cashflow_map
    assert "기타금융자산 변동내역 (단위 : 백만원) · 기타유동금융자산의 취득 · 2,000" in cashflow_map
    assert "이익잉여금 배당 내역 (단위 : 백만원) · 배당금지급 · 1,500" in cashflow_map
    assert "차입금 전기 (단위 : 백만원)" not in cashflow_map
    assert "주석 금액 확인" in cashflow_map
    assert "대사 타깃 미정의" not in cashflow_map
    assert "자동 대사됨" not in cashflow_map


def test_cashflow_map_shows_formula_note_evidence_when_single_amount_is_not_available():
    statements = [
        _section(
            "statement:cfs",
            "현금흐름표",
            "statement",
            "",
            ReportTable(
                0,
                [
                    ["구분", "당기"],
                    ["투자활동으로 인한 현금흐름", ""],
                    ["유형자산의 취득", "(1,800,000,000)"],
                ],
                "현금흐름표",
                SourceLocation("statement:cfs", 0, 0),
            ),
        )
    ]
    checks = [
        CheckResult(
            "cf",
            "cashflow_reconciliation",
            "explainable_gap",
            "note",
            "11",
            "property_plant_equipment.acquisitions_cashflow",
            1_800_000_000,
            1_700_000_000,
            -100_000_000,
            0,
            "주석 취득 2,000,000,000 - 비현금거래-미지급금 증가 300,000,000 = 1,700,000,000; 현금흐름표 유형자산의 취득 1,800,000,000; 차이 (100,000,000); 현금흐름표 금액과 직접 대사되지 않음",
            [
                CheckEvidence("cfs 유형자산의 취득", -1_800_000_000, "statement:cfs/table:0/row:2/col:1"),
                CheckEvidence("note 11 취득", 2_000_000_000, "note:11/table:0/row:1/col:1"),
                CheckEvidence("note 11 비현금거래-미지급금", 300_000_000, "note:11/table:0/row:2/col:1"),
            ],
        )
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", statements, []), checks)
    start = html.index('<section class="report-section report-frame-section" id="cash-flows"')
    cashflow_map = html[start : html.index('<section class="report-section report-frame-section" id="notes"', start)]

    assert "산식 대사 근거" in cashflow_map
    assert "주석 11 취득 · 2,000,000,000" in cashflow_map
    assert "주석 11 비현금거래-미지급금 · 300,000,000" in cashflow_map
    assert "당기 주석에서 동일 금액 미확인" not in cashflow_map


def test_cashflow_map_separates_related_note_difference_from_missing_disclosure():
    statements = [
        _section(
            "statement:cfs",
            "현금흐름표",
            "statement",
            "",
            ReportTable(
                0,
                [
                    ["구분", "당기"],
                    ["투자활동으로 인한 현금흐름", ""],
                    ["유형자산의 취득", "(1,800,000,000)"],
                    ["선급금의 증가", "(1,100,000,000)"],
                ],
                "현금흐름표",
                SourceLocation("statement:cfs", 0, 0),
            ),
        )
    ]
    notes = [
        _section(
            "note:ppe",
            "유형자산",
            "note",
            "11",
            ReportTable(
                1,
                [
                    ["구분", "당기"],
                    ["취득", "1,500"],
                ],
                "유형자산 변동내역 (단위 : 백만원)",
                SourceLocation("note:ppe", 0, 1),
                unit_multiplier=1_000_000,
            ),
        )
    ]

    html = render_audit_reconciliation_html(FullReport("sample.html", "Sample Co", statements, notes), [])
    start = html.index('<section class="report-section report-frame-section" id="cash-flows"')
    cashflow_map = html[start : html.index('<section class="report-section report-frame-section" id="notes"', start)]

    assert "차이내역 확인 필요" in cashflow_map
    assert "관련 주석 확인됨 · 동일 금액 없음" in cashflow_map
    assert "유형자산 변동내역 (단위 : 백만원) · 취득 · 1,500" in cashflow_map
    assert "대응 주석 확인 필요" in cashflow_map
    assert "주석 전수 확인 결과 동일/관련 금액 미확인" in cashflow_map
    assert "추가 확인 필요" not in cashflow_map


def test_html_statement_match_section_includes_self_verification_advisory():
    """재무제표 ↔ 주석 대사 section must surface the self-verification limitation upfront."""
    report = FullReport("sample.html", "Sample Co", [], [])
    html = render_audit_reconciliation_html(report, [])
    assert 'class="self-verify-advisory"' in html
    assert "자체 정합성" in html
    assert "자산 증감표 검산" in html
    # advisory should be a multi-row block (heading + bulleted legend), not one collapsed sentence.
    assert "관련 주석의 자체 검증 한계" in html
    assert "<ul>" in html and "self-verify-badge ok" in html and "self-verify-badge none" in html


def test_html_leadsheet_trigger_uses_vertical_stack_with_meta_row():
    """leadsheet 관련 주석 셀: label 위, badge+상세 메타 row 아래의 vertical stack."""
    table = ReportTable(
        0,
        [["구분", "당기"], ["유형자산", "1,000"]],
        "재무상태표",
        SourceLocation("statement:bs", 0, 0),
    )
    bs_section = ReportSection(
        "statement:bs", "재무상태표", "statement", None,
        [ReportBlock("table", "", table, table.location)],
    )
    note_table = ReportTable(
        0,
        [["구분", "당기"], ["기초장부금액", "900"], ["취득", "100"], ["기말장부금액", "1,000"]],
        "유형자산 변동내역",
        SourceLocation("note:11", 0, 0),
    )
    note_section = ReportSection(
        "note:11", "유형자산 변동내역", "note", "11",
        [ReportBlock("table", "", note_table, note_table.location)],
    )
    report = FullReport("sample.html", "Sample Co", [bs_section], [note_section])
    html = render_audit_reconciliation_html(report, [])
    leadsheet_start = html.index('<table class="leadsheet">')
    leadsheet_end = html.index("</table>", leadsheet_start)
    leadsheet_html = html[leadsheet_start:leadsheet_end]
    assert '<span class="leadsheet-note-display">' in leadsheet_html
    assert '<span class="leadsheet-note-meta">' in leadsheet_html
    display_idx = leadsheet_html.index('class="leadsheet-note-display"')
    meta_idx = leadsheet_html.index('class="leadsheet-note-meta"')
    assert display_idx < meta_idx


def test_html_leadsheet_related_note_cell_uses_note_number_and_hover_trigger(tmp_path):
    """leadsheet 관련 주석 셀: 주석 NN. 주제명 + row-match-trigger + hover-note."""
    current = tmp_path / "current.html"
    current.write_text(
        """
        <p>재무상태표</p><table><tr><th>구분</th><th>당기</th></tr><tr><td>유형자산</td><td>1,000</td></tr></table>
        <p>손익계산서</p><table><tr><th>구분</th><th>당기</th></tr><tr><td>매출액</td><td>2,000</td></tr></table>
        <p>현금흐름표</p><table><tr><th>구분</th><th>당기</th></tr><tr><td>유형자산의 취득</td><td>(500)</td></tr></table>
        <p>11. 유형자산</p><table><tr><th>구분</th><th>당기</th><th>전기</th><th>합계</th></tr><tr><td>기초</td><td>800</td><td>700</td><td>1,500</td></tr><tr><td>취득</td><td>500</td><td>400</td><td>900</td></tr><tr><td>장부금액</td><td>1,000</td><td>800</td><td>1,800</td></tr></table>
        """,
        encoding="utf-8",
    )
    output = tmp_path / "workpaper.html"
    result = CliRunner().invoke(app, ["workpaper-html", str(current), str(output), "--company", "Sample Co"])
    assert result.exit_code == 0
    html = output.read_text(encoding="utf-8")

    leadsheet_start = html.index('<table class="leadsheet">')
    leadsheet_end = html.index("</table>", leadsheet_start)
    leadsheet_html = html[leadsheet_start:leadsheet_end]

    assert "related-note-cell" in leadsheet_html
    assert "leadsheet-note-trigger" in leadsheet_html
    assert "row-match-trigger" in leadsheet_html
    assert "leadsheet-note-display" in leadsheet_html
    assert "주석 11." in leadsheet_html
    assert "유형자산" in leadsheet_html
    assert 'class="hover-note"' in leadsheet_html
    assert "self-verify-badge" in leadsheet_html


def test_html_leadsheet_self_verification_badge_marks_verified_when_rollforward_matched():
    """leadsheet 셀의 self-verify-badge는 note_rollforward_check 결과를 반영."""
    table = ReportTable(
        0,
        [
            ["구분", "당기"],
            ["유형자산", "1,000"],
        ],
        "재무상태표",
        SourceLocation("statement:bs", 0, 0),
    )
    bs_section = ReportSection("statement:bs", "재무상태표", "statement", None, [ReportBlock("table", "", table, table.location)])
    note_table = ReportTable(
        0,
        [
            ["구분", "당기"],
            ["기초장부금액", "900"],
            ["취득", "100"],
            ["기말장부금액", "1,000"],
        ],
        "유형자산",
        SourceLocation("note:11", 0, 0),
    )
    note_section = ReportSection("note:11", "유형자산", "note", "11", [ReportBlock("table", "", note_table, note_table.location)])
    report = FullReport("sample.html", "Sample Co", [bs_section], [note_section])

    matched_rollforward = CheckResult(
        "note_assertion:11:table0:rollforward:col1",
        "note_rollforward_check",
        "matched",
        "note",
        "11",
        "유형자산 증감표 검산 - 당기",
        1_000,
        1_000,
        0,
        0,
        "기초와 변동내역이 기말 장부금액과 일치",
        [
            CheckEvidence("기초장부금액 당기", 900, "note:11/table:0/row:1/col:1"),
            CheckEvidence("기말장부금액 당기", 1_000, "note:11/table:0/row:3/col:1"),
        ],
    )
    html = render_audit_reconciliation_html(report, [matched_rollforward])
    leadsheet_start = html.index('<table class="leadsheet">')
    leadsheet_end = html.index("</table>", leadsheet_start)
    leadsheet_html = html[leadsheet_start:leadsheet_end]
    assert 'class="self-verify-badge ok"' in leadsheet_html
    assert "증감표 검산" in leadsheet_html


def test_html_check_row_wraps_title_with_hover_trigger_when_note_evidence_available():
    """`자산 주석 연결 대사` 등 _check_row 첫 컬럼이 note_no + raw note table이 있으면 hover-trigger로 감싸짐."""
    note_table = ReportTable(
        0,
        [
            ["구분", "당기"],
            ["기초", "100"],
            ["취득", "20"],
            ["기말", "120"],
        ],
        "무형자산 증감표",
        SourceLocation("note:14", 0, 0),
    )
    note_section = ReportSection("note:14", "무형자산", "note", "14", [ReportBlock("table", "", note_table, note_table.location)])
    report = FullReport("sample.html", "Sample Co", [], [note_section])
    bridge_check = CheckResult(
        "asset_note_bridge:intangible_assets.acquisitions_cashflow",
        "asset_note_bridge_check",
        "matched",
        "report",
        "14",
        "무형자산 취득 주석 연결 대사",
        20,
        20,
        0,
        0,
        "자산 주석과 관련 주석 금액이 현금흐름표 산식으로 연결됨",
        [
            CheckEvidence("cfs 무형자산의 취득", -20, "statement:cf/table:0/row:1/col:1"),
            CheckEvidence("note 14 취득", 20, "note:14/table:0/row:2/col:1"),
        ],
    )
    html = render_audit_reconciliation_html(report, [bridge_check])
    bridge_start = html.index('id="notes"')
    bridge_end = html.index("</section>", bridge_start)
    bridge_html = html[bridge_start:bridge_end]
    assert "재무제표-주석 대사" in bridge_html
    assert "무형자산 취득 주석 연결 대사" in bridge_html
    assert "주석 14 · 표 0 · 행 2 · 열 1" in bridge_html
